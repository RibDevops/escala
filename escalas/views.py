"""
Views do Sistema de Escala Militar.
Foco atual: dashboard, autenticação e cadastros (OM, Divisões, Postos,
Especialidades, Militares). As views de geração de escala estão em
`views_escala_legado.py` e serão integradas em uma próxima etapa.

Suporta múltiplas OMs: a OM ativa é mantida na sessão do usuário
(`request.session['om_id_ativa']`) e selecionada via dropdown na navbar.
"""
from datetime import date as _date, date, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .context_processors import SESSION_KEY_OM, obter_om_da_sessao
from .forms_cadastro import (
    DivisaoForm,
    EscalaCriarForm,
    EspecialidadeForm,
    IndisponibilidadeRegistrarForm,
    LancamentoManualForm,
    MilitarForm,
    OrganizacaoMilitarForm,
    PostoForm,
    QuadrinhoForm,
    TipoEscalaForm,
    TipoIndisponibilidadeForm,
    UsuarioForm,
)
from .models import (
    CalendarioDia,
    ConfiguracaoEscala,
    Divisao,
    Escala,
    EscalaCalendarioOverride,
    EscalaItem,
    Especialidade,
    Indisponibilidade,
    LancamentoManualQuadrinho,
    Militar,
    OrganizacaoMilitar,
    PerfilUsuario,
    Posto,
    Quadrinho,
    TipoEscala,
    TipoIndisponibilidade,
    TipoServico,
    TrocaServico,
    UsuarioCustomizado,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def obter_om_ativa(request):
    """Retorna a OM ativa do usuário (sessão) ou None se nenhuma cadastrada."""
    return obter_om_da_sessao(request)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    om = obter_om_ativa(request)

    contexto = {
        'om': om,
        'total_militares': 0,
        'total_divisoes': 0,
        'total_postos': Posto.objects.filter(ativo=True).count(),
        'total_especialidades': Especialidade.objects.filter(ativo=True).count(),
        'total_oms': OrganizacaoMilitar.objects.filter(ativo=True).count(),
        'militares_recentes': [],
        'divisoes_resumo': [],
    }

    if om:
        contexto.update({
            'total_militares': Militar.objects.filter(
                organizacao_militar=om, ativo=True
            ).count(),
            'total_divisoes': Divisao.objects.filter(
                organizacao_militar=om, ativo=True
            ).count(),
            'militares_recentes': Militar.objects.filter(
                organizacao_militar=om, ativo=True
            ).select_related('posto', 'divisao', 'especialidade').order_by(
                '-data_criacao'
            )[:6],
            'divisoes_resumo': Divisao.objects.filter(
                organizacao_militar=om, ativo=True
            ).annotate(
                total_militares=Count('militares', filter=Q(militares__ativo=True))
            ).order_by('nome'),
        })

    return render(request, 'dashboard.html', contexto)


# ---------------------------------------------------------------------------
# Organizações Militares (multi-OM)
# ---------------------------------------------------------------------------

@login_required
def organizacao_listar(request):
    oms = OrganizacaoMilitar.objects.annotate(
        total_militares=Count('militares', filter=Q(militares__ativo=True)),
        total_divisoes=Count('divisoes', filter=Q(divisoes__ativo=True), distinct=True),
    ).order_by('-ativo', 'sigla')
    return render(request, 'cadastro/organizacao_list.html', {'oms': oms})


@login_required
def organizacao_detalhe(request, om_id=None):
    """Detalhe de uma OM. Se om_id não informado, usa a OM ativa."""
    if om_id is None:
        om = obter_om_ativa(request)
        if om is None:
            return redirect('organizacao_novo')
    else:
        om = get_object_or_404(OrganizacaoMilitar, pk=om_id)
    return render(request, 'cadastro/organizacao_detail.html', {'om': om})


@login_required
def organizacao_form(request, om_id=None):
    instancia = get_object_or_404(OrganizacaoMilitar, pk=om_id) if om_id else None
    if request.method == 'POST':
        form = OrganizacaoMilitarForm(request.POST, request.FILES, instance=instancia)
        if form.is_valid():
            om = form.save()
            # se for a primeira, ativa na sessão
            if not request.session.get(SESSION_KEY_OM):
                request.session[SESSION_KEY_OM] = om.id
            messages.success(
                request,
                f'OM {om.sigla} {"atualizada" if instancia else "cadastrada"} com sucesso.',
            )
            return redirect('organizacao_detalhe', om_id=om.id)
    else:
        form = OrganizacaoMilitarForm(instance=instancia)
    return render(
        request,
        'cadastro/organizacao_form.html',
        {'form': form, 'om': instancia},
    )


@login_required
def organizacao_excluir(request, om_id):
    om = get_object_or_404(OrganizacaoMilitar, pk=om_id)
    qtd_militares = om.militares.count()
    qtd_escalas = om.escalas.count()
    qtd_divisoes = om.divisoes.count()
    tem_vinculos = bool(qtd_militares or qtd_escalas or qtd_divisoes)
    if request.method == 'POST':
        om_ativa_id = request.session.get(SESSION_KEY_OM)
        if tem_vinculos:
            om.ativo = False
            om.save()
            if om_ativa_id == om.id:
                request.session.pop(SESSION_KEY_OM, None)
            messages.success(
                request,
                f'OM {om.sigla} desativada (existem militares, escalas ou divisões '
                'vinculados; histórico preservado).',
            )
        else:
            if om_ativa_id == om.id:
                request.session.pop(SESSION_KEY_OM, None)
            om.delete()
            messages.success(request, f'OM excluída.')
        return redirect('organizacao_listar')
    return render(
        request,
        'cadastro/organizacao_confirm_delete.html',
        {
            'om': om,
            'qtd_militares': qtd_militares,
            'qtd_escalas': qtd_escalas,
            'qtd_divisoes': qtd_divisoes,
            'tem_vinculos': tem_vinculos,
        },
    )


@login_required
@require_POST
def organizacao_trocar(request):
    """Define a OM ativa na sessão e volta para a página anterior."""
    om_id_raw = request.POST.get('om_id')
    proximo = request.POST.get('next') or 'dashboard'
    try:
        om_id = int(om_id_raw) if om_id_raw not in (None, '', 'None') else None
    except (TypeError, ValueError):
        om_id = None
    if om_id:
        om = OrganizacaoMilitar.objects.filter(pk=om_id, ativo=True).first()
        if om:
            request.session[SESSION_KEY_OM] = om.id
            messages.success(request, f'OM ativa alterada para {om.sigla}.')
        else:
            messages.error(request, 'OM inválida ou inativa.')
    return redirect(proximo)


# ---------------------------------------------------------------------------
# Postos (lista global, hierarquia militar)
# ---------------------------------------------------------------------------

@login_required
def posto_listar(request):
    postos = Posto.objects.all().order_by('ordem_hierarquica')
    return render(request, 'cadastro/posto_list.html', {'postos': postos})


@login_required
def posto_form(request, posto_id=None):
    instancia = get_object_or_404(Posto, pk=posto_id) if posto_id else None
    if request.method == 'POST':
        form = PostoForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Posto salvo com sucesso.')
            return redirect('posto_listar')
    else:
        form = PostoForm(instance=instancia)
    return render(
        request,
        'cadastro/posto_form.html',
        {'form': form, 'posto': instancia},
    )


@login_required
def posto_excluir(request, posto_id):
    posto = get_object_or_404(Posto, pk=posto_id)
    if request.method == 'POST':
        if posto.militares.filter(ativo=True).exists():
            messages.error(
                request,
                'Existem militares ativos com este posto. '
                'Desative o posto ao invés de excluir.',
            )
        else:
            posto.ativo = False
            posto.save()
            messages.success(request, 'Posto desativado.')
        return redirect('posto_listar')
    return render(
        request,
        'cadastro/posto_confirm_delete.html',
        {'posto': posto},
    )


# ---------------------------------------------------------------------------
# Especialidades
# ---------------------------------------------------------------------------

@login_required
def especialidade_listar(request):
    especialidades = Especialidade.objects.all().order_by('nome')
    return render(
        request,
        'cadastro/especialidade_list.html',
        {'especialidades': especialidades},
    )


@login_required
def especialidade_form(request, especialidade_id=None):
    instancia = (
        get_object_or_404(Especialidade, pk=especialidade_id)
        if especialidade_id else None
    )
    if request.method == 'POST':
        form = EspecialidadeForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Especialidade salva com sucesso.')
            return redirect('especialidade_listar')
    else:
        form = EspecialidadeForm(instance=instancia)
    return render(
        request,
        'cadastro/especialidade_form.html',
        {'form': form, 'especialidade': instancia},
    )


@login_required
def especialidade_excluir(request, especialidade_id):
    esp = get_object_or_404(Especialidade, pk=especialidade_id)
    if request.method == 'POST':
        esp.ativo = False
        esp.save()
        messages.success(request, 'Especialidade desativada.')
        return redirect('especialidade_listar')
    return render(
        request,
        'cadastro/especialidade_confirm_delete.html',
        {'especialidade': esp},
    )


# ---------------------------------------------------------------------------
# Tipos de Indisponibilidade (lista global)
# ---------------------------------------------------------------------------

@login_required
def tipo_indisponibilidade_listar(request):
    tipos = TipoIndisponibilidade.objects.all().order_by('nome')
    return render(
        request,
        'cadastro/tipo_indisponibilidade_list.html',
        {'tipos': tipos},
    )


@login_required
def tipo_indisponibilidade_form(request, tipo_id=None):
    instancia = (
        get_object_or_404(TipoIndisponibilidade, pk=tipo_id) if tipo_id else None
    )
    if request.method == 'POST':
        form = TipoIndisponibilidadeForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tipo de indisponibilidade salvo com sucesso.')
            return redirect('tipo_indisponibilidade_listar')
    else:
        form = TipoIndisponibilidadeForm(instance=instancia)
    return render(
        request,
        'cadastro/tipo_indisponibilidade_form.html',
        {'form': form, 'tipo': instancia},
    )


@login_required
def tipo_indisponibilidade_excluir(request, tipo_id):
    tipo = get_object_or_404(TipoIndisponibilidade, pk=tipo_id)
    if request.method == 'POST':
        if tipo.indisponibilidades.exists():
            tipo.ativo = False
            tipo.save()
            messages.success(
                request,
                'Tipo de indisponibilidade desativado (existem registros vinculados, '
                'histórico preservado).',
            )
        else:
            tipo.delete()
            messages.success(request, 'Tipo de indisponibilidade excluído.')
        return redirect('tipo_indisponibilidade_listar')
    return render(
        request,
        'cadastro/tipo_indisponibilidade_confirm_delete.html',
        {'tipo': tipo},
    )


# ---------------------------------------------------------------------------
# Tipos de Escala (cadastro global, não escopado por OM)
# ---------------------------------------------------------------------------

@login_required
def tipo_escala_listar(request):
    tipos = (
        TipoEscala.objects.all()
        .annotate(
            qtd_escalas=Count('escalas', distinct=True),
            qtd_quadrinhos=Count('quadrinhos', distinct=True),
        )
        .order_by('-ativo', 'nome')
    )
    return render(
        request,
        'cadastro/tipo_escala_list.html',
        {'tipos': tipos},
    )


@login_required
def tipo_escala_form(request, tipo_id=None):
    instancia = get_object_or_404(TipoEscala, pk=tipo_id) if tipo_id else None
    if request.method == 'POST':
        form = TipoEscalaForm(request.POST, instance=instancia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tipo de escala salvo com sucesso.')
            return redirect('tipo_escala_listar')
    else:
        form = TipoEscalaForm(instance=instancia)
    return render(
        request,
        'cadastro/tipo_escala_form.html',
        {'form': form, 'tipo': instancia},
    )


@login_required
def tipo_escala_excluir(request, tipo_id):
    tipo = get_object_or_404(TipoEscala, pk=tipo_id)
    tem_vinculos = tipo.escalas.exists() or tipo.quadrinhos.exists()
    if request.method == 'POST':
        if tem_vinculos:
            tipo.ativo = False
            tipo.save()
            messages.success(
                request,
                'Tipo de escala desativado (existem escalas ou quadrinhos '
                'vinculados, histórico preservado).',
            )
        else:
            tipo.delete()
            messages.success(request, 'Tipo de escala excluído.')
        return redirect('tipo_escala_listar')
    return render(
        request,
        'cadastro/tipo_escala_confirm_delete.html',
        {
            'tipo': tipo,
            'qtd_escalas': tipo.escalas.count(),
            'qtd_quadrinhos': tipo.quadrinhos.count(),
            'tem_vinculos': tem_vinculos,
        },
    )


# ---------------------------------------------------------------------------
# Divisões (escopo: OM ativa)
# ---------------------------------------------------------------------------

@login_required
def divisao_listar(request):
    om = obter_om_ativa(request)
    divisoes = (
        Divisao.objects.filter(organizacao_militar=om).annotate(
            total_militares=Count('militares', filter=Q(militares__ativo=True))
        ).order_by('nome')
        if om else Divisao.objects.none()
    )
    return render(
        request,
        'cadastro/divisao_list.html',
        {'divisoes': divisoes, 'om': om},
    )


@login_required
def divisao_form(request, divisao_id=None):
    om = obter_om_ativa(request)
    if om is None:
        messages.error(request, 'Cadastre uma Organização Militar antes.')
        return redirect('organizacao_novo')

    instancia = get_object_or_404(Divisao, pk=divisao_id) if divisao_id else None

    if request.method == 'POST':
        form = DivisaoForm(request.POST, instance=instancia)
        if form.is_valid():
            divisao = form.save(commit=False)
            divisao.organizacao_militar = om
            divisao.save()
            messages.success(request, 'Divisão salva com sucesso.')
            return redirect('divisao_listar')
    else:
        form = DivisaoForm(instance=instancia)

    return render(
        request,
        'cadastro/divisao_form.html',
        {'form': form, 'divisao': instancia, 'om': om},
    )


@login_required
def divisao_excluir(request, divisao_id):
    divisao = get_object_or_404(Divisao, pk=divisao_id)
    if request.method == 'POST':
        divisao.ativo = False
        divisao.save()
        messages.success(request, 'Divisão desativada.')
        return redirect('divisao_listar')
    return render(
        request,
        'cadastro/divisao_confirm_delete.html',
        {'divisao': divisao},
    )


# ---------------------------------------------------------------------------
# Militares (escopo: OM ativa)
# ---------------------------------------------------------------------------



@login_required
def militar_detalhe(request, militar_id):
    militar = get_object_or_404(
        Militar.objects.select_related(
            'posto', 'divisao', 'especialidade', 'organizacao_militar'
        ),
        pk=militar_id,
    )

    om = militar.organizacao_militar
    ano_atual = _date.today().year
    try:
        ano = int(request.GET.get('ano') or ano_atual)
    except ValueError:
        ano = ano_atual

    tipos_servico = list(om.tipos_servico.filter(ativo=True).order_by('ordem'))
    tipos_escala = list(TipoEscala.objects.filter(ativo=True).order_by('nome'))

    quadrinhos = Quadrinho.objects.filter(militar=militar, ano=ano).select_related(
        'tipo_escala', 'tipo_servico'
    )
    quadrinhos_map = {(q.tipo_escala_id, q.tipo_servico_id): q for q in quadrinhos}

    contadores = []
    for te in tipos_escala:
        celulas = []
        total = 0
        tem_dado = False
        for ts in tipos_servico:
            qd = quadrinhos_map.get((te.id, ts.id))
            valor = qd.total if qd else 0
            celulas.append({
                'tipo_servico': ts,
                'valor': valor,
                'quadrinho': qd,
            })
            total += valor
            if qd:
                tem_dado = True
        contadores.append({
            'tipo_escala': te,
            'celulas': celulas,
            'total': total,
            'tem_dado': tem_dado,
        })

    itens_qs = (
        EscalaItem.objects.filter(
            militar=militar,
            calendario_dia__data__year=ano,
        )
        .select_related(
            'escala__tipo_escala',
            'calendario_dia__tipo_servico',
        )
        .order_by('calendario_dia__data')
    )
    itens = list(itens_qs)

    dias_servico = {it.calendario_dia.data: it for it in itens}

    DIAS_SEMANA_ABREV = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']

    import calendar as _cal
    cal = _cal.Calendar(firstweekday=0)
    meses = []
    for mes_num in range(1, 13):
        nome_mes = _cal.month_name[mes_num].capitalize() if False else [
            'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
        ][mes_num - 1]
        semanas = []
        tem_servico_no_mes = False
        for semana in cal.monthdatescalendar(ano, mes_num):
            dias_semana = []
            for d in semana:
                pertence_mes = d.month == mes_num
                item = dias_servico.get(d) if pertence_mes else None
                if item:
                    tem_servico_no_mes = True
                dias_semana.append({
                    'data': d,
                    'pertence_mes': pertence_mes,
                    'item': item,
                })
            semanas.append(dias_semana)
        meses.append({
            'numero': mes_num,
            'nome': nome_mes,
            'semanas': semanas,
            'tem_servico': tem_servico_no_mes,
        })

    anos_opcoes = list(range(ano_atual + 1, ano_atual - 5, -1))

    return render(
        request,
        'cadastro/militar_detail.html',
        {
            'militar': militar,
            'ano': ano,
            'anos_opcoes': anos_opcoes,
            'contadores': contadores,
            'tipos_servico': tipos_servico,
            'itens': itens,
            'meses': meses,
            'dias_semana_abrev': DIAS_SEMANA_ABREV,
            'total_servicos_ano': len(itens),
        },
    )


# ---------------------------------------------------------------------------
# Quadrinho (visão geral por OM × Tipo de Escala × Ano)
# ---------------------------------------------------------------------------

@login_required
def quadrinho_visao(request):
    om = obter_om_ativa(request)
    militar_proprio = getattr(request.user, 'militar', None)

    ano_atual = _date.today().year
    try:
        ano = int(request.GET.get('ano') or 0)
    except ValueError:
        ano = 0
    todos_anos = (ano == 0)

    tipo_escala_param = request.GET.get('tipo_escala', '')

    tipos_escala = list(TipoEscala.objects.filter(ativo=True).order_by('nome'))
    tipo_escala_atual = None
    if tipo_escala_param:
        tipo_escala_atual = next(
            (t for t in tipos_escala if str(t.id) == tipo_escala_param), None
        )
    if tipo_escala_atual is None and tipos_escala:
        tipo_escala_atual = tipos_escala[0]

    tipos_servico = (
        list(om.tipos_servico.filter(ativo=True).order_by('ordem')) if om else []
    )

    # Se for militar próprio, filtra somente ele no quadrinho
    if militar_proprio:
        militares = [militar_proprio] if om else []
    else:
        militares = (
            list(
                Militar.objects.filter(organizacao_militar=om, ativo=True)
                .select_related('posto', 'divisao')
            )
            if om else []
        )

    # ── Carregar quadrinhos (todos os anos ou ano específico) ─────────────────
    # quadrinhos_info_map: (mil_id, ts_id) -> {'ajuste': int, 'quantidade': int, 'qd': obj|None}
    quadrinhos_info_map = {}
    if om and tipo_escala_atual and militares and tipos_servico:
        from django.db.models import Sum as _Sum
        filtro_q = dict(
            militar__in=militares,
            tipo_escala=tipo_escala_atual,
            tipo_servico__in=tipos_servico,
        )
        if not todos_anos:
            filtro_q['ano'] = ano

        if todos_anos:
            for row in Quadrinho.objects.filter(**filtro_q).values(
                'militar_id', 'tipo_servico_id'
            ).annotate(total_aj=_Sum('ajuste_inicial'), total_qt=_Sum('quantidade')):
                quadrinhos_info_map[(row['militar_id'], row['tipo_servico_id'])] = {
                    'ajuste_inicial': row['total_aj'] or 0,
                    'quantidade': row['total_qt'] or 0,
                    'qd': None,
                }
        else:
            for qd in Quadrinho.objects.filter(**filtro_q):
                quadrinhos_info_map[(qd.militar_id, qd.tipo_servico_id)] = {
                    'ajuste_inicial': qd.ajuste_inicial,
                    'quantidade': qd.quantidade,
                    'qd': qd,
                }

    # Lançamentos manuais somados por (militar, tipo_servico)
    lancamentos_totais_map = {}  # (mil_id, ts_id) -> soma de quantidades
    if om and tipo_escala_atual and militares and tipos_servico:
        filtro_lm = dict(
            militar__in=militares,
            tipo_escala=tipo_escala_atual,
            tipo_servico__in=tipos_servico,
        )
        if not todos_anos:
            filtro_lm['ano'] = ano
        for lm in LancamentoManualQuadrinho.objects.filter(**filtro_lm):
            chave = (lm.militar_id, lm.tipo_servico_id)
            lancamentos_totais_map[chave] = lancamentos_totais_map.get(chave, 0) + lm.quantidade

    linhas = []
    totais_coluna = {ts.id: 0 for ts in tipos_servico}
    total_geral = 0
    for m in militares:
        celulas = []
        total_militar = 0
        for ts in tipos_servico:
            info = quadrinhos_info_map.get((m.id, ts.id), {'ajuste_inicial': 0, 'quantidade': 0, 'qd': None})
            valor_sistema = info['ajuste_inicial'] + info['quantidade']
            valor_manual = lancamentos_totais_map.get((m.id, ts.id), 0)
            valor = valor_sistema + valor_manual
            celulas.append({
                'tipo_servico': ts,
                'quadrinho': info['qd'],
                'valor': valor,
                'valor_sistema': valor_sistema,
                'valor_manual': valor_manual,
                'ajuste_inicial': info['ajuste_inicial'],
                'quantidade': info['quantidade'],
            })
            totais_coluna[ts.id] += valor
            total_militar += valor
        linhas.append({
            'militar': m,
            'celulas': celulas,
            'total': total_militar,
        })
        total_geral += total_militar

    ordem = request.GET.get('ordem', 'desc')
    if ordem == 'asc':
        linhas.sort(key=lambda x: (x['total'], x['militar'].nome_guerra))
    elif ordem == 'nome':
        linhas.sort(key=lambda x: x['militar'].nome_guerra.lower())
    elif ordem == 'antiguidade':
        from datetime import date as _date_cls
        def _ant_key(x):
            m = x['militar']
            data = m.data_ultima_promocao if m.data_ultima_promocao else _date_cls.max
            nota = -float(m.nota) if m.nota is not None else float('inf')
            return (m.posto.ordem_hierarquica, data, nota, m.nome_guerra.lower())
        linhas.sort(key=_ant_key)
    else:
        linhas.sort(key=lambda x: (-x['total'], x['militar'].nome_guerra))

    totais_coluna_lista = [
        {'tipo_servico': ts, 'valor': totais_coluna[ts.id]} for ts in tipos_servico
    ]

    anos_opcoes = list(range(ano_atual + 1, ano_atual - 5, -1))

    # Escala atual do mês/ano para exibir links de Matriz e Detalhe no quadrinho
    from datetime import date as _d
    hoje = _d.today()
    escala_atual = None
    if om and tipo_escala_atual:
        escala_atual = (
            Escala.objects.filter(
                organizacao_militar=om,
                tipo_escala=tipo_escala_atual,
                ano=ano,
                mes=hoje.month,
            )
            .order_by('-data_criacao')
            .first()
        )
        if escala_atual is None:
            escala_atual = (
                Escala.objects.filter(
                    organizacao_militar=om,
                    tipo_escala=tipo_escala_atual,
                    ano=ano,
                )
                .order_by('-mes', '-data_criacao')
                .first()
            )

    # ── Registros detalhados de serviços por militar ───────────────────────
    # Regra: se há substituto, o serviço é contabilizado para ELE (não o titular).
    registros_por_militar = {}
    if om and tipo_escala_atual and militares:
        from django.db.models import Q as _Q
        militar_ids = [m.id for m in militares]
        militar_ids_set = set(militar_ids)

        # Busca:
        #  (a) itens onde este militar é titular E não há substituto (serviu de fato)
        #  (b) itens onde este militar é o substituto (cobriu outro)
        filtro_itens_reg = dict(escala__tipo_escala=tipo_escala_atual)
        if not todos_anos:
            filtro_itens_reg['calendario_dia__data__year'] = ano
        itens = EscalaItem.objects.filter(
            _Q(militar_id__in=militar_ids, substituto__isnull=True) |
            _Q(substituto_id__in=militar_ids),
            **filtro_itens_reg,
        ).select_related(
            'militar__posto',
            'substituto__posto',
            'calendario_dia__tipo_servico',
            'escala',
        ).order_by('calendario_dia__data')

        for item in itens:
            # Quem de fato trabalhou neste dia?
            if item.substituto_id and item.substituto_id in militar_ids_set:
                mil_id = item.substituto_id
            else:
                mil_id = item.militar_id

            if mil_id not in registros_por_militar:
                registros_por_militar[mil_id] = []
            registros_por_militar[mil_id].append({
                'data': item.calendario_dia.data,
                'tipo_servico': item.calendario_dia.tipo_servico,
                'escala_mes': f"{item.escala.mes:02d}/{item.escala.ano}",
                'observacao': item.observacao or '',
                'id': item.id,
            })

    # ── Matriz: militares × dias registrados, separado por tipo de serviço ──
    # Filtro de período da matriz (separado dos filtros de totais)
    try:
        matriz_mes = int(request.GET.get('matriz_mes') or 0)  # 0 = todos os meses
    except ValueError:
        matriz_mes = 0
    try:
        matriz_ano = int(request.GET.get('matriz_ano') or 0)  # 0 = todos os anos
    except ValueError:
        matriz_ano = 0
    todos_anos_matriz = (matriz_ano == 0)

    # Militares em ordem de antiguidade para a Matriz (independente do filtro de totais)
    militares_antiguidade = (
        list(
            Militar.objects.filter(organizacao_militar=om, ativo=True)
            .select_related('posto')
            .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
        )
        if om else []
    )

    # Lançamentos manuais da matriz, agrupados por (mil_id, ts_id)
    lancamentos_matriz = {}  # (mil_id, ts_id) -> list of LancamentoManualQuadrinho
    if om and tipo_escala_atual and militares_antiguidade and tipos_servico:
        filtro_lm_mat = dict(
            militar_id__in=[m.id for m in militares_antiguidade],
            tipo_escala=tipo_escala_atual,
            tipo_servico_id__in=[ts.id for ts in tipos_servico],
        )
        if not todos_anos_matriz:
            filtro_lm_mat['ano'] = matriz_ano
        for lm in LancamentoManualQuadrinho.objects.filter(**filtro_lm_mat).select_related('tipo_servico'):
            chave = (lm.militar_id, lm.tipo_servico_id)
            lancamentos_matriz.setdefault(chave, []).append(lm)

    # Seções por tipo de serviço
    matriz_secoes = []
    if om and tipo_escala_atual and militares_antiguidade and tipos_servico:
        filtro_itens = dict(
            militar_id__in=[m.id for m in militares_antiguidade],
            escala__tipo_escala=tipo_escala_atual,
        )
        if not todos_anos_matriz:
            filtro_itens['calendario_dia__data__year'] = matriz_ano
        if matriz_mes:
            filtro_itens['calendario_dia__data__month'] = matriz_mes

        todos_itens = (
            EscalaItem.objects.filter(**filtro_itens)
            .values(
                'id',
                'militar_id',
                'observacao',
                'calendario_dia__data',
                'calendario_dia__tipo_servico_id',
                'calendario_dia__tipo_servico__nome',
                'calendario_dia__tipo_servico__cor_hex',
            )
            .order_by('calendario_dia__data')
        )

        # Agrupar por tipo de serviço → {ts_id: {mil_id: [(data, item_id, obs)]}}
        por_ts = {}
        for item in todos_itens:
            ts_id = item['calendario_dia__tipo_servico_id']
            mil_id = item['militar_id']
            if ts_id not in por_ts:
                por_ts[ts_id] = {
                    'nome': item['calendario_dia__tipo_servico__nome'],
                    'cor': item['calendario_dia__tipo_servico__cor_hex'],
                    'por_militar': {},
                }
            por_ts[ts_id]['por_militar'].setdefault(mil_id, []).append(
                (item['calendario_dia__data'], item['id'], item['observacao'] or '')
            )

        # Ajuste inicial por (mil_id, ts_id) — histórico pré-sistema sem datas
        # Somado por todos os anos quando todos_anos_matriz=True
        from django.db.models import Sum as _SumQ
        ajuste_mat_filtro = dict(
            militar_id__in=[m.id for m in militares_antiguidade],
            tipo_escala=tipo_escala_atual,
            tipo_servico_id__in=[ts.id for ts in tipos_servico],
        )
        if not todos_anos_matriz:
            ajuste_mat_filtro['ano'] = matriz_ano
        ajuste_inicial_map = {}  # (mil_id, ts_id) -> total ajuste_inicial
        for row in Quadrinho.objects.filter(**ajuste_mat_filtro).values(
            'militar_id', 'tipo_servico_id'
        ).annotate(total_aj=_SumQ('ajuste_inicial')):
            ajuste_inicial_map[(row['militar_id'], row['tipo_servico_id'])] = row['total_aj'] or 0

        for ts in tipos_servico:
            bloco = por_ts.get(ts.id)

            linhas_ts = []
            total_geral_ts = 0
            max_entradas = 0

            for m in militares_antiguidade:
                # Entradas de histórico pré-sistema (ajuste_inicial sem datas)
                n_historico = ajuste_inicial_map.get((m.id, ts.id), 0)
                entradas_historico = [{'tipo': 'historico'} for _ in range(n_historico)]

                # Lançamentos manuais expandidos (cada lm com qtd=N vira N células)
                lms = lancamentos_matriz.get((m.id, ts.id), [])
                entradas_manuais = []
                for lm in lms:
                    for _ in range(lm.quantidade):
                        entradas_manuais.append({
                            'tipo': 'manual',
                            'label': lm.label,
                            'cat': lm.tipo,
                            'id': lm.id,
                            'observacao': lm.observacao or '',
                        })

                # Serviços reais do sistema em ordem cronológica
                raw_reais = sorted(bloco['por_militar'].get(m.id, []), key=lambda x: x[0]) if bloco else []
                entradas_reais = [
                    {'tipo': 'data', 'data': d, 'id': eid, 'observacao': obs}
                    for d, eid, obs in raw_reais
                ]

                # Ordem: histórico pré-sistema → serviços reais → lançamentos manuais
                entradas = entradas_historico + entradas_reais + entradas_manuais
                total = len(entradas)
                total_geral_ts += total
                max_entradas = max(max_entradas, total)

                linhas_ts.append({
                    'militar': m,
                    'entradas': entradas,
                    'total': total,
                })

            colunas = list(range(1, max_entradas + 1))

            matriz_secoes.append({
                'tipo_servico': ts,
                'max_entradas': max_entradas,
                'colunas': colunas,
                'linhas': linhas_ts,
                'total_geral': total_geral_ts,
            })

    anos_opcoes_matriz = [0] + list(range(ano_atual + 1, ano_atual - 5, -1))

    return render(
        request,
        'cadastro/quadrinho_visao.html',
        {
            'om': om,
            'ano': ano,
            'todos_anos': todos_anos,
            'anos_opcoes': anos_opcoes,
            'tipos_escala': tipos_escala,
            'tipo_escala_atual': tipo_escala_atual,
            'tipos_servico': tipos_servico,
            'linhas': linhas,
            'totais_coluna': totais_coluna_lista,
            'total_geral': total_geral,
            'ordem': ordem,
            'escala_atual': escala_atual,
            'militar_proprio': militar_proprio,
            'registros_por_militar': registros_por_militar,
            'matriz_secoes': matriz_secoes,
            'matriz_mes': matriz_mes,
            'matriz_ano': matriz_ano,
            'todos_anos_matriz': todos_anos_matriz,
            'anos_opcoes_matriz': anos_opcoes_matriz,
            'nomes_meses': NOMES_MESES,
        },
    )


# ---------------------------------------------------------------------------
# Indisponibilidades — auto-serviço do militar e gestão pelo escalante
# ---------------------------------------------------------------------------

@login_required
def indisponibilidade_listar(request):
    om = obter_om_ativa(request)
    militar_proprio = getattr(request.user, 'militar', None)

    if militar_proprio:
        indisp = (
            Indisponibilidade.objects.filter(militar=militar_proprio)
            .select_related('tipo', 'militar__posto')
            .order_by('-data_inicio')
        )
        militares = None
        filtro_mil = None
    else:
        indisp = (
            Indisponibilidade.objects.filter(militar__organizacao_militar=om)
            .select_related('tipo', 'militar__posto')
            .order_by('-data_inicio')
        )
        militares = (
            Militar.objects.filter(organizacao_militar=om, ativo=True)
            .select_related('posto')
            .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
        ) if om else []
        filtro_mil = request.GET.get('militar', '')
        if filtro_mil:
            indisp = indisp.filter(militar_id=filtro_mil)


    return render(request, 'indisponibilidade/listar.html', {
        'indisp': indisp,
        'om': om,
        'militar_proprio': militar_proprio,
        'militares': militares,
        'filtro_mil': filtro_mil,
    })


@login_required
def indisponibilidade_criar(request):
    om = obter_om_ativa(request)
    militar_proprio = getattr(request.user, 'militar', None)

    if request.method == 'POST':
        form = IndisponibilidadeRegistrarForm(
            request.POST, om=om, militar_fixo=militar_proprio
        )
        if form.is_valid():
            ind = form.save(commit=False)
            if militar_proprio and not ind.militar_id:
                ind.militar = militar_proprio
            ind.data_fim = form.cleaned_data['data_fim']
            ind.save()
            um_dia = form.cleaned_data['data_inicio'] == form.cleaned_data['data_fim']
            if getattr(form, '_data_fim_ajustada', False):
                msg = (
                    f'Indisponibilidade registrada como 1 dia: '
                    f'{ind.tipo.nome} em {ind.data_inicio.strftime("%d/%m/%Y")}.'
                )
            else:
                msg = (
                    f'Indisponibilidade registrada: {ind.tipo.nome} de '
                    f'{ind.data_inicio.strftime("%d/%m/%Y")} a {ind.data_fim.strftime("%d/%m/%Y")}.'
                )
            messages.success(request, msg)
            return redirect('indisponibilidade_listar')
    else:
        form = IndisponibilidadeRegistrarForm(om=om, militar_fixo=militar_proprio)

    return render(request, 'indisponibilidade/criar.html', {
        'form': form,
        'om': om,
        'militar_proprio': militar_proprio,
    })


@login_required
@require_POST
def indisponibilidade_excluir(request, ind_id):
    ind = get_object_or_404(Indisponibilidade, pk=ind_id)
    militar_proprio = getattr(request.user, 'militar', None)

    if militar_proprio and ind.militar_id != militar_proprio.id:
        messages.error(request, 'Sem permissão para excluir esta indisponibilidade.')
        return redirect('indisponibilidade_listar')

    desc = f'{ind.tipo.nome} — {ind.militar.nome_guerra}'
    ind.delete()
    messages.success(request, f'Indisponibilidade removida: {desc}.')
    return redirect('indisponibilidade_listar')


@login_required
def quadrinho_editar(request, militar_id, tipo_escala_id, tipo_servico_id, ano):
    om = obter_om_ativa(request)
    militar = get_object_or_404(
        Militar.objects.select_related('posto', 'organizacao_militar'),
        pk=militar_id,
    )
    if om and militar.organizacao_militar_id != om.id:
        messages.error(request, 'O militar não pertence à OM ativa.')
        return redirect('quadrinho_visao')

    tipo_escala = get_object_or_404(TipoEscala, pk=tipo_escala_id)
    tipo_servico = get_object_or_404(
        TipoServico, pk=tipo_servico_id, organizacao_militar=militar.organizacao_militar
    )

    quadrinho, _ = Quadrinho.objects.get_or_create(
        militar=militar,
        tipo_escala=tipo_escala,
        tipo_servico=tipo_servico,
        ano=ano,
        defaults={'quantidade': 0, 'ajuste_inicial': 0},
    )

    voltar_para = request.GET.get('voltar', 'quadrinho_visao')

    if request.method == 'POST':
        form = QuadrinhoForm(request.POST, instance=quadrinho)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f'Quadrinho de {militar.nome_guerra} atualizado '
                f'({tipo_escala.nome} / {tipo_servico.nome}).'
            )
            if voltar_para == 'militar_detalhe':
                return redirect(
                    f"{reverse('militar_detalhe', args=[militar.id])}?ano={ano}"
                )
            return redirect(
                f"{reverse('quadrinho_visao')}?ano={ano}"
                f"&tipo_escala={tipo_escala.id}"
            )
    else:
        form = QuadrinhoForm(instance=quadrinho)

    return render(
        request,
        'cadastro/quadrinho_form.html',
        {
            'form': form,
            'quadrinho': quadrinho,
            'militar': militar,
            'tipo_escala': tipo_escala,
            'tipo_servico': tipo_servico,
            'ano': ano,
            'voltar_para': voltar_para,
        },
    )


# ---------------------------------------------------------------------------
# Quadrinho — views AJAX para modais (editar, excluir, add lançamento)
# ---------------------------------------------------------------------------

@login_required
def quadrinho_ajax_editar(request, militar_id, tipo_escala_id, tipo_servico_id, ano):
    militar = get_object_or_404(Militar.objects.select_related('posto'), pk=militar_id)
    tipo_escala = get_object_or_404(TipoEscala, pk=tipo_escala_id)
    tipo_servico = get_object_or_404(TipoServico, pk=tipo_servico_id)

    quadrinho, _ = Quadrinho.objects.get_or_create(
        militar=militar,
        tipo_escala=tipo_escala,
        tipo_servico=tipo_servico,
        ano=ano,
        defaults={'quantidade': 0, 'ajuste_inicial': 0},
    )

    def _valor_manual():
        return sum(
            lm.quantidade
            for lm in LancamentoManualQuadrinho.objects.filter(
                militar=militar, tipo_escala=tipo_escala,
                tipo_servico=tipo_servico, ano=ano,
            )
        )

    if request.method == 'POST':
        try:
            ajuste = max(0, int(request.POST.get('ajuste_inicial', 0)))
            quantidade = max(0, int(request.POST.get('quantidade', 0)))
        except (ValueError, TypeError):
            return JsonResponse({'ok': False, 'erro': 'Valores inválidos.'}, status=400)
        quadrinho.ajuste_inicial = ajuste
        quadrinho.quantidade = quantidade
        quadrinho.save()
        vm = _valor_manual()
        return JsonResponse({
            'ok': True,
            'ajuste_inicial': quadrinho.ajuste_inicial,
            'quantidade': quadrinho.quantidade,
            'valor_sistema': quadrinho.total,
            'valor_manual': vm,
            'total': quadrinho.total + vm,
        })

    vm = _valor_manual()
    return JsonResponse({
        'ok': True,
        'militar_nome': f'{militar.posto.sigla} {militar.nome_guerra}',
        'tipo_servico_nome': tipo_servico.nome,
        'ajuste_inicial': quadrinho.ajuste_inicial,
        'quantidade': quadrinho.quantidade,
        'valor_sistema': quadrinho.total,
        'valor_manual': vm,
        'total': quadrinho.total + vm,
    })


@login_required
@require_POST
def quadrinho_ajax_excluir(request, militar_id, tipo_escala_id, tipo_servico_id, ano):
    militar = get_object_or_404(Militar, pk=militar_id)
    tipo_escala = get_object_or_404(TipoEscala, pk=tipo_escala_id)
    tipo_servico = get_object_or_404(TipoServico, pk=tipo_servico_id)
    Quadrinho.objects.filter(
        militar=militar, tipo_escala=tipo_escala,
        tipo_servico=tipo_servico, ano=ano,
    ).update(ajuste_inicial=0, quantidade=0)
    vm = sum(
        lm.quantidade
        for lm in LancamentoManualQuadrinho.objects.filter(
            militar=militar, tipo_escala=tipo_escala,
            tipo_servico=tipo_servico, ano=ano,
        )
    )
    return JsonResponse({'ok': True, 'total': vm, 'valor_sistema': 0, 'valor_manual': vm})


@login_required
@require_POST
def lancamento_ajax_criar(request):
    try:
        militar_id = int(request.POST['militar_id'])
        tipo_escala_id = int(request.POST['tipo_escala_id'])
        tipo_servico_id = int(request.POST['tipo_servico_id'])
        ano = int(request.POST['ano'])
        tipo = request.POST.get('tipo', 'lastro')
        label = request.POST.get('label', '').strip()
        quantidade = max(1, int(request.POST.get('quantidade', 1)))
        observacao = request.POST.get('observacao', '')
    except (KeyError, ValueError, TypeError):
        return JsonResponse({'ok': False, 'erro': 'Dados inválidos.'}, status=400)

    if not label:
        return JsonResponse({'ok': False, 'erro': 'A descrição é obrigatória.'}, status=400)

    militar = get_object_or_404(Militar, pk=militar_id)
    tipo_escala = get_object_or_404(TipoEscala, pk=tipo_escala_id)
    tipo_servico = get_object_or_404(TipoServico, pk=tipo_servico_id)

    LancamentoManualQuadrinho.objects.create(
        militar=militar,
        tipo_escala=tipo_escala,
        tipo_servico=tipo_servico,
        ano=ano,
        tipo=tipo,
        label=label,
        quantidade=quantidade,
        observacao=observacao,
        criado_por=request.user,
    )

    qd = Quadrinho.objects.filter(
        militar=militar, tipo_escala=tipo_escala,
        tipo_servico=tipo_servico, ano=ano,
    ).first()
    valor_sistema = qd.total if qd else 0
    valor_manual = sum(
        lm.quantidade
        for lm in LancamentoManualQuadrinho.objects.filter(
            militar=militar, tipo_escala=tipo_escala,
            tipo_servico=tipo_servico, ano=ano,
        )
    )
    return JsonResponse({
        'ok': True,
        'valor_sistema': valor_sistema,
        'valor_manual': valor_manual,
        'total': valor_sistema + valor_manual,
    })


# ---------------------------------------------------------------------------
# Quadrinho Matriz — AJAX para editar/excluir lançamento e EscalaItem
# ---------------------------------------------------------------------------

@login_required
def lancamento_ajax_editar_por_id(request, lancamento_id):
    lm = get_object_or_404(
        LancamentoManualQuadrinho.objects.select_related('militar__posto', 'tipo_servico'),
        pk=lancamento_id,
    )
    if request.method == 'GET':
        return JsonResponse({
            'ok': True,
            'tipo': lm.tipo,
            'label': lm.label,
            'quantidade': lm.quantidade,
            'observacao': lm.observacao or '',
            'militar_nome': f'{lm.militar.posto.sigla} {lm.militar.nome_guerra}',
            'tipo_servico_nome': lm.tipo_servico.nome,
        })
    # POST
    try:
        tipo = request.POST.get('tipo', lm.tipo)
        label = request.POST.get('label', '').strip()
        quantidade = max(1, int(request.POST.get('quantidade', 1)))
        observacao = request.POST.get('observacao', '')
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'erro': 'Dados inválidos.'}, status=400)
    if not label:
        return JsonResponse({'ok': False, 'erro': 'A descrição é obrigatória.'}, status=400)
    lm.tipo = tipo
    lm.label = label
    lm.quantidade = quantidade
    lm.observacao = observacao
    lm.save()
    return JsonResponse({'ok': True, 'reload': True})


@login_required
@require_POST
def lancamento_ajax_excluir_por_id(request, lancamento_id):
    lm = get_object_or_404(LancamentoManualQuadrinho, pk=lancamento_id)
    if lm.quantidade > 1:
        lm.quantidade -= 1
        lm.save(update_fields=['quantidade'])
    else:
        lm.delete()
    return JsonResponse({'ok': True, 'reload': True})


@login_required
def escala_item_ajax_editar(request, item_id):
    item = get_object_or_404(
        EscalaItem.objects.select_related(
            'militar__posto', 'calendario_dia__tipo_servico'
        ),
        pk=item_id,
    )
    if request.method == 'GET':
        return JsonResponse({
            'ok': True,
            'data': item.calendario_dia.data.strftime('%d/%m/%Y'),
            'observacao': item.observacao or '',
            'militar_nome': f'{item.militar.posto.sigla} {item.militar.nome_guerra}',
            'tipo_servico_nome': item.calendario_dia.tipo_servico.nome,
        })
    # POST
    item.observacao = request.POST.get('observacao', '').strip()
    item.save()
    return JsonResponse({'ok': True, 'reload': True})


@login_required
@require_POST
def escala_item_ajax_excluir(request, item_id):
    item = get_object_or_404(
        EscalaItem.objects.select_related(
            'militar', 'escala__tipo_escala', 'calendario_dia__tipo_servico'
        ),
        pk=item_id,
    )
    militar = item.militar
    tipo_escala = item.escala.tipo_escala
    tipo_servico = item.calendario_dia.tipo_servico
    ano = item.calendario_dia.data.year
    item.delete()
    # Recalcula quantidade no quadrinho após remoção
    new_count = EscalaItem.objects.filter(
        militar=militar,
        escala__tipo_escala=tipo_escala,
        calendario_dia__tipo_servico=tipo_servico,
        calendario_dia__data__year=ano,
    ).count()
    Quadrinho.objects.filter(
        militar=militar, tipo_escala=tipo_escala,
        tipo_servico=tipo_servico, ano=ano,
    ).update(quantidade=new_count)
    return JsonResponse({'ok': True, 'reload': True})


@login_required
def militar_form(request, militar_id=None):
    om = obter_om_ativa(request)
    if om is None:
        messages.error(request, 'Cadastre uma Organização Militar antes.')
        return redirect('organizacao_novo')

    instancia = get_object_or_404(Militar, pk=militar_id) if militar_id else None

    if request.method == 'POST':
        form = MilitarForm(request.POST, instance=instancia, om=om)
        if form.is_valid():
            militar = form.save(commit=False)
            militar.organizacao_militar = om
            militar.save()
            form.save_m2m()  # salva tipos_escala

            # Criar usuário automaticamente se ainda não tem um
            if not militar.user_id:
                try:
                    usuario = militar.criar_usuario_automatico()
                    messages.info(
                        request,
                        f'Usuário <strong>{usuario.username}</strong> criado automaticamente. '
                        f'Senha inicial: matrícula ({militar.matricula}).',
                    )
                except Exception as exc:
                    messages.warning(
                        request,
                        f'Militar salvo, mas não foi possível criar o usuário automático: {exc}',
                    )
            else:
                # Atualiza campos de acesso do usuário vinculado
                usuario = militar.user
                novo_username = request.POST.get('user_username', '').strip().lower()
                novo_perfil = request.POST.get('user_perfil', '')
                user_ativo = request.POST.get('user_ativo') == 'on'
                changed = False
                if novo_username and novo_username != usuario.username:
                    if not UsuarioCustomizado.objects.filter(username=novo_username).exclude(pk=usuario.pk).exists():
                        usuario.username = novo_username
                        changed = True
                    else:
                        messages.warning(request, f'Username "{novo_username}" já está em uso; mantido o anterior.')
                if novo_perfil and novo_perfil in dict(PerfilUsuario.choices) and usuario.perfil != novo_perfil:
                    usuario.perfil = novo_perfil
                    changed = True
                if usuario.ativo != user_ativo or usuario.is_active != user_ativo:
                    usuario.ativo = user_ativo
                    usuario.is_active = user_ativo
                    changed = True
                if changed:
                    usuario.save()

            messages.success(
                request,
                f'Militar {militar.nome_guerra} salvo com sucesso.',
            )
            return redirect('militar_detalhe', militar_id=militar.id)
    else:
        form = MilitarForm(instance=instancia, om=om)

    return render(
        request,
        'cadastro/militar_form.html',
        {'form': form, 'militar': instancia, 'om': om, 'perfis': PerfilUsuario.choices},
    )


@login_required
def militar_excluir(request, militar_id):
    militar = get_object_or_404(Militar, pk=militar_id)
    if request.method == 'POST':
        militar.ativo = False
        militar.save()
        messages.success(
            request,
            f'Militar {militar.nome_guerra} desativado (histórico preservado).',
        )
        return redirect('usuario_listar')
    return render(
        request,
        'cadastro/militar_confirm_delete.html',
        {'militar': militar},
    )


# ===========================================================================
# ESCALAS — listagem, criação, detalhe, geração automática (matriz)
# ===========================================================================

NOMES_MESES = [
    '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
]


@login_required
def escala_listar(request):
    """Lista todas as escalas da OM ativa."""
    om = obter_om_ativa(request)
    if not om:
        messages.warning(request, 'Nenhuma OM ativa. Cadastre ou selecione uma OM.')
        return redirect('organizacao_novo')

    escalas = (
        Escala.objects.filter(organizacao_militar=om)
        .select_related('tipo_escala')
        .order_by('-ano', '-mes')
    )

    # filtros simples
    filtro_tipo = request.GET.get('tipo')
    filtro_status = request.GET.get('status')
    filtro_ano = request.GET.get('ano')
    if filtro_tipo:
        escalas = escalas.filter(tipo_escala_id=filtro_tipo)
    if filtro_status:
        escalas = escalas.filter(status=filtro_status)
    if filtro_ano:
        escalas = escalas.filter(ano=filtro_ano)

    tipos = TipoEscala.objects.filter(ativo=True)
    anos = sorted(
        Escala.objects.filter(organizacao_militar=om)
        .values_list('ano', flat=True).distinct(),
        reverse=True,
    )

    return render(request, 'escala/listar.html', {
        'escalas': escalas,
        'tipos': tipos,
        'anos': anos,
        'filtro_tipo': filtro_tipo,
        'filtro_status': filtro_status,
        'filtro_ano': filtro_ano,
        'STATUS_CHOICES': Escala.STATUS_CHOICES,
        'nomes_meses': NOMES_MESES,
    })


@login_required
def escala_criar(request):
    """Cria um novo cabeçalho de escala para a OM ativa."""
    om = obter_om_ativa(request)
    if not om:
        messages.warning(request, 'Nenhuma OM ativa.')
        return redirect('organizacao_novo')

    if request.method == 'POST':
        form = EscalaCriarForm(request.POST)
        if form.is_valid():
            tipo = form.cleaned_data['tipo_escala']
            mes = int(form.cleaned_data['mes'])
            ano = form.cleaned_data['ano']
            obs = form.cleaned_data.get('observacao', '')

            # Verificar duplicidade
            if Escala.objects.filter(
                organizacao_militar=om, tipo_escala=tipo, mes=mes, ano=ano
            ).exists():
                messages.error(
                    request,
                    f'Já existe uma escala de {tipo.nome} para '
                    f'{NOMES_MESES[mes]}/{ano}.',
                )
            else:
                escala = Escala.objects.create(
                    organizacao_militar=om,
                    tipo_escala=tipo,
                    mes=mes,
                    ano=ano,
                    observacao=obs,
                    status='previsao',
                )
                messages.success(
                    request,
                    f'Escala {NOMES_MESES[mes]}/{ano} criada com sucesso!',
                )
                return redirect('escala_detalhar', escala_id=escala.id)
    else:
        form = EscalaCriarForm()

    return render(request, 'escala/criar.html', {'form': form, 'om': om})


@login_required
def escala_detalhar(request, escala_id):
    """Exibe os itens da escala em formato de lista e tabela."""
    om = obter_om_ativa(request)
    escala = get_object_or_404(Escala, pk=escala_id)

    itens = (
        escala.itens
        .select_related('militar__posto', 'calendario_dia__tipo_servico', 'substituto__posto')
        .order_by('calendario_dia__data')
    )

    # Contagem por militar
    contagem: dict = {}
    for item in itens:
        m = item.substituto if item.substituto else item.militar
        contagem.setdefault(m, 0)
        contagem[m] += 1

    contagem_lista = sorted(contagem.items(), key=lambda x: -x[1])

    # Lista de militares ativos da OM para os selects
    militares_om = list(
        Militar.objects.filter(
            organizacao_militar=escala.organizacao_militar, ativo=True
        )
        .select_related('posto')
        .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
    )

    return render(request, 'escala/detalhar.html', {
        'escala': escala,
        'itens': itens,
        'contagem_lista': contagem_lista,
        'nomes_meses': NOMES_MESES,
        'pode_editar': escala.status == 'previsao',
        'militares_om': militares_om,
    })


@login_required
@require_POST
def escala_item_trocar_militar(request, item_id):
    """
    Rascunho/Previsão: troca o militar de um item.
    Atualiza o quadrinho: decrementa o militar anterior e incrementa o novo.
    Chamado via AJAX (fetch) com body JSON {'militar_id': <id>}.
    """
    from django.http import JsonResponse
    import json

    item = get_object_or_404(EscalaItem, pk=item_id)
    escala = item.escala

    if escala.status != 'previsao':
        return JsonResponse({'ok': False, 'erro': 'Escala não editável.'}, status=400)

    try:
        data = json.loads(request.body)
        novo_militar_id = int(data.get('militar_id', 0))
    except (ValueError, KeyError):
        return JsonResponse({'ok': False, 'erro': 'Dados inválidos.'}, status=400)

    if novo_militar_id == item.militar_id:
        return JsonResponse({'ok': True, 'mensagem': 'Sem alteração.'})

    novo_militar = get_object_or_404(
        Militar, pk=novo_militar_id, organizacao_militar=escala.organizacao_militar, ativo=True
    )

    militar_anterior = item.militar
    tipo_escala = escala.tipo_escala
    tipo_servico = item.calendario_dia.tipo_servico
    ano = escala.ano

    # Decrementa quadrinho do militar anterior
    try:
        qd_anterior = Quadrinho.objects.get(
            militar=militar_anterior,
            tipo_escala=tipo_escala,
            tipo_servico=tipo_servico,
            ano=ano,
        )
        if qd_anterior.quantidade > 0:
            qd_anterior.quantidade -= 1
            qd_anterior.save(update_fields=['quantidade'])
    except Quadrinho.DoesNotExist:
        pass

    # Incrementa quadrinho do novo militar
    Quadrinho.incrementar(
        militar=novo_militar,
        tipo_escala=tipo_escala,
        tipo_servico=tipo_servico,
        ano=ano,
    )

    # Atualiza o item — bypass da validação de indisponibilidade via update_fields
    item.militar = novo_militar
    item.save(update_fields=['militar'])

    return JsonResponse({
        'ok': True,
        'nome_guerra': novo_militar.nome_guerra,
        'posto_sigla': novo_militar.posto.sigla,
    })


@login_required
@require_POST
def escala_item_definir_substituto(request, item_id):
    """
    Escala Oficial: define o substituto de um item.
    Transfere o quadrinho: decrementa o titular, incrementa o substituto.
    Remove substituto se militar_id='' for enviado.
    """
    from django.http import JsonResponse
    import json

    item = get_object_or_404(
        EscalaItem.objects.select_related(
            'militar__posto', 'substituto__posto',
            'calendario_dia__tipo_servico', 'escala__tipo_escala',
        ),
        pk=item_id,
    )
    escala = item.escala

    if escala.status != 'publicada':
        return JsonResponse({'ok': False, 'erro': 'Apenas em Escala Oficial.'}, status=400)

    try:
        data = json.loads(request.body)
        novo_sub_id = data.get('militar_id', '')
    except (ValueError, KeyError):
        return JsonResponse({'ok': False, 'erro': 'Dados inválidos.'}, status=400)

    tipo_escala = escala.tipo_escala
    tipo_servico = item.calendario_dia.tipo_servico
    ano = escala.ano

    substituto_anterior = item.substituto

    # --- Remover substituto ---
    if novo_sub_id == '' or novo_sub_id is None:
        if substituto_anterior:
            # Devolve quadrinho ao titular original
            try:
                qd_sub = Quadrinho.objects.get(
                    militar=substituto_anterior,
                    tipo_escala=tipo_escala,
                    tipo_servico=tipo_servico,
                    ano=ano,
                )
                if qd_sub.quantidade > 0:
                    qd_sub.quantidade -= 1
                    qd_sub.save(update_fields=['quantidade'])
            except Quadrinho.DoesNotExist:
                pass
            Quadrinho.incrementar(
                militar=item.militar,
                tipo_escala=tipo_escala,
                tipo_servico=tipo_servico,
                ano=ano,
            )
            item.substituto = None
            item.save(update_fields=['substituto'])
        return JsonResponse({'ok': True, 'removido': True})

    try:
        novo_sub_id = int(novo_sub_id)
    except ValueError:
        return JsonResponse({'ok': False, 'erro': 'ID inválido.'}, status=400)

    if novo_sub_id == item.militar_id:
        return JsonResponse({'ok': False, 'erro': 'Substituto não pode ser o próprio titular.'}, status=400)

    novo_sub = get_object_or_404(
        Militar, pk=novo_sub_id, organizacao_militar=escala.organizacao_militar, ativo=True
    )

    # Se já havia substituto anterior diferente do novo: só reverte o sub antigo.
    # O titular NÃO recebe +1 aqui — ele continua "fora" do serviço (conta -1).
    # A contagem do titular só volta quando o substituto é REMOVIDO (bloco acima).
    if substituto_anterior and substituto_anterior.pk != novo_sub.pk:
        try:
            qd_sub_ant = Quadrinho.objects.get(
                militar=substituto_anterior,
                tipo_escala=tipo_escala,
                tipo_servico=tipo_servico,
                ano=ano,
            )
            if qd_sub_ant.quantidade > 0:
                qd_sub_ant.quantidade -= 1
                qd_sub_ant.save(update_fields=['quantidade'])
        except Quadrinho.DoesNotExist:
            pass
        # Não devolve ao titular — ele ainda está "substituído" neste dia.

    # Se é substituição nova (não havia sub antes), decrementa o titular pela 1ª vez
    if not substituto_anterior:
        try:
            qd_titular = Quadrinho.objects.get(
                militar=item.militar,
                tipo_escala=tipo_escala,
                tipo_servico=tipo_servico,
                ano=ano,
            )
            if qd_titular.quantidade > 0:
                qd_titular.quantidade -= 1
                qd_titular.save(update_fields=['quantidade'])
        except Quadrinho.DoesNotExist:
            pass

    # Incrementa quadrinho do novo substituto
    Quadrinho.incrementar(
        militar=novo_sub,
        tipo_escala=tipo_escala,
        tipo_servico=tipo_servico,
        ano=ano,
    )

    item.substituto = novo_sub
    item.save(update_fields=['substituto'])

    return JsonResponse({
        'ok': True,
        'nome_guerra': novo_sub.nome_guerra,
        'posto_sigla': novo_sub.posto.sigla,
    })


@login_required
def escala_gerar(request, escala_id):
    """Redireciona para o novo motor vertical (único algoritmo ativo)."""
    return redirect('escala_gerar_vertical', escala_id=escala_id)


@login_required
def escala_gerar_vertical(request, escala_id):
    """
    Gera a escala usando o MotorEscalaVertical (único algoritmo ativo).

    Algoritmo:
      - Menor quantidade total de serviços → prioridade
      - Desempate: BASE → TOPO (mais moderno primeiro)
      - Folga GLOBAL entre Preto/Vermelho (configurável em horas)
      - Preto gerado primeiro, depois Vermelho
      - Dois estados: Matriz Histórica (banco) + Matriz Operacional (memória)

    GET  = tela de confirmação (ou exibe resultado da última geração via session)
    POST = executa o motor e redireciona para GET com resultado na session
    """
    from django.core.exceptions import ValidationError
    from .services import gerar_escala_vertical as _gerar

    escala = get_object_or_404(Escala, pk=escala_id)
    om = escala.organizacao_militar

    if escala.status != 'previsao':
        messages.error(request, 'Somente escalas em Previsão podem ser geradas.')
        return redirect('escala_detalhar', escala_id=escala_id)

    if request.method == 'POST':
        try:
            resultado = _gerar(escala)
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('escala_detalhar', escala_id=escala_id)

        # Guardar resultado na session para exibir no GET
        request.session['motor_resultado'] = {
            'alocacoes_criadas': resultado['alocacoes_criadas'],
            'dias_sem_militar':  resultado['dias_sem_militar'],
            'alertas':           resultado['alertas'],
            'log':               resultado['log'],
            'escala_id':         escala_id,
        }

        if resultado['dias_sem_militar']:
            messages.warning(
                request,
                f"{resultado['dias_sem_militar']} dia(s) sem cobertura — "
                f"todos os militares estavam com férias/licença."
            )
        if resultado['alertas']:
            messages.warning(
                request,
                f"{len(resultado['alertas'])} alerta(s) gerado(s) durante a geração."
            )
        messages.success(
            request,
            f"Escala gerada! {resultado['alocacoes_criadas']} serviço(s) atribuído(s)."
        )
        return redirect('escala_gerar_vertical', escala_id=escala_id)

    # GET — tela de confirmação ou resultado
    resultado = None
    session_resultado = request.session.pop('motor_resultado', None)
    if session_resultado and session_resultado.get('escala_id') == escala_id:
        resultado = session_resultado

    militares_count = Militar.objects.filter(organizacao_militar=om, ativo=True).count()
    tem_itens = escala.itens.exists()

    # Calendário do mês para visualização/edição na tela de gerar
    import calendar as _cal
    primeiro_dia = date(escala.ano, escala.mes, 1)
    ultimo_num = _cal.monthrange(escala.ano, escala.mes)[1]
    ultimo_dia = date(escala.ano, escala.mes, ultimo_num)

    dias_om = {
        cd.data: cd
        for cd in CalendarioDia.objects.filter(
            organizacao_militar=om,
            data__range=(primeiro_dia, ultimo_dia),
        ).select_related('tipo_servico')
    }
    overrides = {
        ov.data: ov
        for ov in EscalaCalendarioOverride.objects.filter(
            escala=escala,
        ).select_related('tipo_servico')
    }

    tipos_servico_om = list(
        TipoServico.objects.filter(organizacao_militar=om, ativo=True).order_by('ordem')
    )

    dias_calendario = []
    for n in range(1, ultimo_num + 1):
        d = date(escala.ano, escala.mes, n)
        cd = dias_om.get(d)
        ov = overrides.get(d)
        tipo_atual = ov.tipo_servico if ov else (cd.tipo_servico if cd else None)
        dias_calendario.append({
            'data': d,
            'tipo': tipo_atual,
            'override': ov is not None,
            'dia_semana': d.strftime('%a'),
        })

    # Verifica permissão de edição do calendário (Chefe, Adjunto ou Escalante)
    pode_editar_calendario = request.user.perfil in (
        PerfilUsuario.CHEFE, PerfilUsuario.ADJUNTO, PerfilUsuario.ESCALANTE, PerfilUsuario.ADMIN_OM,
    )

    return render(request, 'escala/gerar_vertical.html', {
        'escala':                   escala,
        'militares_count':          militares_count,
        'tem_itens':                tem_itens,
        'nomes_meses':              NOMES_MESES,
        'resultado':                resultado,
        'dias_calendario':          dias_calendario,
        'tipos_servico_om':         tipos_servico_om,
        'pode_editar_calendario':   pode_editar_calendario,
    })


@login_required
def escala_calendario_trocar_tipo(request, escala_id):
    """
    AJAX — Troca o tipo de serviço de um dia específico dentro desta escala.

    POST  : { data: 'YYYY-MM-DD', tipo_servico_id: <int> }
    DELETE: { data: 'YYYY-MM-DD' }  → remove o override (volta ao padrão da OM)

    Permissão: Chefe, Adjunto, Escalante ou Admin OM.
    Escopo: POR ESCALA — não altera o calendário global da OM.

    Efeito colateral: se a escala já tem itens gerados no dia alterado,
    o EscalaItem existente tem seu CalendarioDia trocado para o novo tipo.
    """
    import json

    escala = get_object_or_404(Escala, pk=escala_id)

    if escala.status != 'previsao':
        return JsonResponse({'ok': False, 'erro': 'Apenas Previsão.'}, status=400)

    perfis_permitidos = (
        PerfilUsuario.CHEFE, PerfilUsuario.ADJUNTO,
        PerfilUsuario.ESCALANTE, PerfilUsuario.ADMIN_OM,
    )
    if request.user.perfil not in perfis_permitidos:
        return JsonResponse({'ok': False, 'erro': 'Sem permissão.'}, status=403)

    try:
        body = json.loads(request.body)
        data_str = body.get('data', '')
        d = date.fromisoformat(data_str)
    except (ValueError, KeyError):
        return JsonResponse({'ok': False, 'erro': 'Data inválida.'}, status=400)

    om = escala.organizacao_militar

    # ── REMOVER override (volta ao padrão da OM) ─────────────────────────────
    if request.method == 'DELETE' or body.get('remover'):
        EscalaCalendarioOverride.objects.filter(escala=escala, data=d).delete()

        # Reverte itens existentes neste dia para o CalendarioDia padrão da OM
        cd_padrao = CalendarioDia.objects.filter(organizacao_militar=om, data=d).first()
        if cd_padrao:
            EscalaItem.objects.filter(
                escala=escala,
                calendario_dia__data=d,
            ).update(calendario_dia=cd_padrao)

        tipo_padrao = cd_padrao.tipo_servico if cd_padrao else None
        return JsonResponse({
            'ok': True,
            'removido': True,
            'tipo_nome': tipo_padrao.nome if tipo_padrao else '',
            'tipo_cor': tipo_padrao.cor_hex if tipo_padrao else '#888',
        })

    # ── APLICAR override ─────────────────────────────────────────────────────
    try:
        tipo_servico_id = int(body.get('tipo_servico_id', 0))
        novo_tipo = TipoServico.objects.get(pk=tipo_servico_id, organizacao_militar=om, ativo=True)
    except (ValueError, TipoServico.DoesNotExist):
        return JsonResponse({'ok': False, 'erro': 'Tipo de serviço inválido.'}, status=400)

    # Obtém ou cria o CalendarioDia para o novo tipo (pode não existir se for
    # uma data que a OM ainda não tem no calendário — improvável mas seguro)
    cd_novo, _ = CalendarioDia.objects.get_or_create(
        organizacao_militar=om,
        data=d,
        defaults={'tipo_servico': novo_tipo, 'origem_tipo': 'MANUAL'},
    )
    # Se já existia com outro tipo, precisamos de um CalendarioDia com o novo tipo.
    # Como unique_together=(om, data) existe, usamos o override para mapear.
    # Os itens apontarão para cd_novo apenas se o tipo coincidir.
    # Estratégia: atualizar o CalendarioDia existente se origem=AUTO e o tipo muda.
    if cd_novo.tipo_servico != novo_tipo:
        # Precisa encontrar ou criar um cd com esse tipo; como unique_together impede
        # dois cd para o mesmo (om, data), atualizamos apenas se for AUTO.
        if cd_novo.origem_tipo == 'AUTO':
            cd_novo.tipo_servico = novo_tipo
            cd_novo.origem_tipo = 'MANUAL'
            cd_novo.save(update_fields=['tipo_servico', 'origem_tipo'])

    # Salva o override para rastrear a mudança por escala
    EscalaCalendarioOverride.objects.update_or_create(
        escala=escala,
        data=d,
        defaults={
            'tipo_servico': novo_tipo,
            'criado_por': request.user,
        },
    )

    # Atualiza itens existentes neste dia para o novo CalendarioDia
    EscalaItem.objects.filter(
        escala=escala,
        calendario_dia__data=d,
    ).update(calendario_dia=cd_novo)

    return JsonResponse({
        'ok': True,
        'tipo_nome': novo_tipo.nome,
        'tipo_cor': novo_tipo.cor_hex,
        'override': True,
    })


@login_required
@require_POST
def escala_limpar(request, escala_id):
    """Remove todos os itens da escala (só Previsão)."""
    escala = get_object_or_404(Escala, pk=escala_id)
    if escala.status != 'previsao':
        messages.error(request, 'Não é possível limpar uma escala publicada.')
    else:
        total = escala.itens.count()
        escala.itens.all().delete()
        messages.success(request, f'{total} item(ns) removido(s).')
    return redirect('escala_detalhar', escala_id=escala_id)


@login_required
@require_POST
def escala_item_forcar(request, item_id):
    """Toggle do flag forcar_escala em um EscalaItem."""
    item = get_object_or_404(EscalaItem, pk=item_id)
    escala = item.escala
    if escala.status == 'previsao':
        item.forcar_escala = not item.forcar_escala
        item.save(update_fields=['forcar_escala'])
        if item.forcar_escala:
            messages.warning(
                request,
                f'{item.militar.nome_guerra} em {item.calendario_dia.data:%d/%m/%Y} '
                'marcado como FORÇADO — folga mínima ignorada para este dia.',
            )
        else:
            messages.success(
                request,
                f'Exceção removida de {item.militar.nome_guerra} '
                f'em {item.calendario_dia.data:%d/%m/%Y}.',
            )
    else:
        messages.error(request, 'Não é possível alterar itens de uma escala publicada.')
    return redirect('escala_detalhar', escala_id=escala.id)



@login_required
@require_POST
def escala_publicar(request, escala_id):
    """Publica a escala (status → publicada). Apenas Chefe ou Adjunto podem publicar."""
    usuario = request.user
    if not usuario.pode_publicar_escala():
        messages.error(request, 'Apenas Chefes ou Adjuntos podem publicar escalas.')
        return redirect('escala_detalhar', escala_id=escala_id)

    escala = get_object_or_404(Escala, pk=escala_id)
    if not escala.itens.exists():
        messages.error(request, 'Escala vazia — preencha antes de publicar.')
        return redirect('escala_detalhar', escala_id=escala_id)
    try:
        escala.publicar()
        messages.success(request, 'Escala publicada com sucesso!')
    except Exception as e:
        messages.error(request, str(e))
    return redirect('escala_detalhar', escala_id=escala_id)


@login_required
@require_POST
def escala_excluir(request, escala_id):
    """Exclui completamente uma escala (cabeçalho + todos os itens)."""
    escala = get_object_or_404(Escala, pk=escala_id)
    if escala.status == 'publicada':
        messages.error(request, 'Escalas publicadas não podem ser excluídas.')
        return redirect('escala_detalhar', escala_id=escala_id)
    nome = str(escala)
    escala.delete()
    messages.success(request, f'Escala "{nome}" excluída com sucesso.')
    return redirect('escala_listar')


@login_required
def configuracao_escala(request):
    """Tela de configuração das regras operacionais da escala para a OM ativa."""
    om = obter_om_ativa(request)
    if not om:
        messages.warning(request, 'Selecione ou cadastre uma OM antes de configurar.')
        return redirect('organizacao_novo')

    config = ConfiguracaoEscala.obter_para_om(om)

    if request.method == 'POST':
        try:
            folga = int(request.POST.get('folga_minima_horas', 48))
            duracao = int(request.POST.get('duracao_servico_horas', 24))
            if folga < 0 or duracao < 1:
                raise ValueError
        except (TypeError, ValueError):
            messages.error(request, 'Valores inválidos. Informe números inteiros positivos.')
            return redirect('configuracao_escala')

        config.folga_minima_horas = folga
        config.duracao_servico_horas = duracao
        config.bloquear_pre_ferias = 'bloquear_pre_ferias' in request.POST
        config.bloquear_pos_ferias = 'bloquear_pos_ferias' in request.POST
        config.save()
        messages.success(request, 'Configurações salvas com sucesso.')
        return redirect('configuracao_escala')

    return render(request, 'escala/configuracao.html', {'config': config, 'om': om})


@login_required
def escala_matriz(request, escala_id):
    """Visualização da matriz algoritmo: militares × dias + passo a passo."""
    escala = get_object_or_404(Escala, pk=escala_id)
    om = escala.organizacao_militar  # usa a OM da própria escala (igual ao escala_detalhar)

    # Militares ordenados ASC: índice 0 = mais antigo (topo), último = mais moderno (base)
    militares = list(
        Militar.objects.filter(organizacao_militar=om, ativo=True)
        .select_related('posto')
        .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
    )

    # Dias do calendário do mês
    dias = list(
        CalendarioDia.objects.filter(
            organizacao_militar=om,
            data__year=escala.ano,
            data__month=escala.mes,
        ).select_related('tipo_servico').order_by('data')
    )

    # Mapa de itens salvos: data → militar
    itens = list(
        escala.itens.select_related('militar__posto', 'calendario_dia__tipo_servico')
        .order_by('calendario_dia__data')
    )
    itens_map = {item.calendario_dia.data: item.militar for item in itens}

    # Indisponibilidades do período
    import calendar as _cal
    ultimo_dia_num = _cal.monthrange(escala.ano, escala.mes)[1]
    inicio = date(escala.ano, escala.mes, 1)
    fim = date(escala.ano, escala.mes, ultimo_dia_num)

    indisp_map = {}  # (militar_id, data) → motivo (str)
    for ind in (
        Indisponibilidade.objects.filter(
            militar__organizacao_militar=om,
            data_inicio__lte=fim,
            data_fim__gte=inicio,
        )
        .select_related('militar', 'tipo')
    ):
        d = ind.data_inicio
        while d <= ind.data_fim:
            if inicio <= d <= fim:
                indisp_map[(ind.militar_id, d)] = ind.tipo.nome
            d += timedelta(days=1)

    # ── Construir linhas da matriz ──────────────────────────────────────
    # Tabela visual: topo = mais antigo (ordem menor), base = mais moderno (ordem maior)
    # Iteramos militares em ordem ASC (mais antigo primeiro = topo da tabela HTML)
    # e invertemos para exibir de baixo para cima no template via CSS flex-direction:column-reverse

    matrix_rows = []
    for mil in militares:
        cells = []
        total_servicos = 0
        for dia in dias:
            d = dia.data
            serves = itens_map.get(d) == mil
            unavailable = (mil.id, d) in indisp_map
            motivo_indisp = indisp_map.get((mil.id, d), '')
            if serves:
                total_servicos += 1
            cells.append({
                'dia': dia,
                'serves': serves,
                'unavailable': unavailable,
                'motivo': motivo_indisp,
            })
        # eventos: só dias em que serviu ou estava indisponível (para colunas da tabela)
        eventos = [c for c in cells if c['serves'] or c['unavailable']]
        matrix_rows.append({
            'militar': mil,
            'cells': cells,
            'eventos': eventos,
            'total': total_servicos,
        })

    # ── Passo a passo: reconstruir raciocínio por dia ───────────────────
    contagem = {mil.id: 0 for mil in militares}
    ultimo_serv = {mil.id: None for mil in militares}
    passos = []

    for dia in dias:
        d = dia.data
        escolhido = itens_map.get(d)

        candidatos = []
        indisponiveis = []
        for mil in militares:
            if (mil.id, d) in indisp_map:
                indisponiveis.append({'militar': mil, 'motivo': indisp_map[(mil.id, d)]})
            else:
                ult = ultimo_serv[mil.id]
                dias_desde = (d - ult).days if ult else None
                candidatos.append({
                    'militar': mil,
                    'count': contagem[mil.id],
                    'dias_desde': dias_desde,
                    'escolhido': mil == escolhido,
                })

        # Ordenar candidatos pela mesma lógica do engine (para exibição)
        candidatos.sort(key=lambda c: (
            c['count'],
            -(c['dias_desde'] if c['dias_desde'] is not None else 9999),
            -militares.index(c['militar']),  # base→topo
        ))

        # Atualizar acumuladores APÓS montar o passo
        if escolhido:
            contagem[escolhido.id] += 1
            ultimo_serv[escolhido.id] = d

        passos.append({
            'dia': dia,
            'escolhido': escolhido,
            'candidatos': candidatos,
            'indisponiveis': indisponiveis,
        })

    max_eventos = max((len(r['eventos']) for r in matrix_rows), default=0)
    # Padear cada linha com None até max_eventos para facilitar o template
    for r in matrix_rows:
        faltam = max_eventos - len(r['eventos'])
        r['eventos_padded'] = r['eventos'] + [None] * faltam

    return render(request, 'escala/matriz.html', {
        'escala': escala,
        'militares': militares,
        'dias': dias,
        'matrix_rows': matrix_rows,
        'max_eventos': max_eventos,
        'max_eventos_range': range(max_eventos),
        'passos': passos,
        'itens': itens,
        'nomes_meses': NOMES_MESES,
    })


# ---------------------------------------------------------------------------
# Documentação do Motor de Escala
# ---------------------------------------------------------------------------

@login_required
def documentacao_motor(request):
    """Página de documentação técnica do motor de geração de escalas."""
    return render(request, 'documentacao/motor_escala.html')


# ---------------------------------------------------------------------------
# Tela pública — Escala de Sobreaviso
# ---------------------------------------------------------------------------

def sobreaviso_publico(request):
    """Redireciona para a nova página pública do primeiro tipo de escala ativo."""
    te = TipoEscala.objects.filter(ativo=True).order_by('nome').first()
    if te and te.slug:
        return redirect('escala_publica', slug=te.slug)
    return redirect('/')


def quadrinho_publico(request):
    """Redireciona para a nova página pública de matriz do primeiro tipo de escala ativo."""
    te = TipoEscala.objects.filter(ativo=True).order_by('nome').first()
    if te and te.slug:
        return redirect('matriz_publica', slug=te.slug)
    return redirect('/')


# ---------------------------------------------------------------------------
# Páginas públicas dinâmicas por slug de TipoEscala
# ---------------------------------------------------------------------------

def escala_publica_redirect(request, slug):
    """Redireciona URLs legadas /escala/<slug>/ para a nova /escala-do-mes/<slug>/."""
    return redirect('escala_publica', slug=slug)


def escala_publica(request, slug):
    """
    Página pública (sem login) de escala para um tipo específico.
    URL: /escala-do-mes/<slug>/
    URL com navegação: /escala-do-mes/<slug>/?mes=5&ano=2026

    Mostra todos os dias do mês selecionado. Suporta navegação entre meses.
    """
    import calendar as _cal
    from datetime import date as _d, timedelta as _td

    hoje = _d.today()
    DIAS_SEMANA = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
    NOMES_MESES_PT = [
        '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
    ]

    tipo_escala = get_object_or_404(TipoEscala, slug=slug, ativo=True)
    om = OrganizacaoMilitar.objects.filter(ativo=True).order_by('id').first()

    todos_tipos = list(TipoEscala.objects.filter(ativo=True).order_by('nome'))
    tipos_servico = list(om.tipos_servico.filter(ativo=True).order_by('ordem')) if om else []

    # Parâmetros de navegação
    try:
        req_mes = int(request.GET.get('mes', 0))
        req_ano = int(request.GET.get('ano', 0))
    except (ValueError, TypeError):
        req_mes = req_ano = 0

    # Todas as escalas disponíveis para navegação (mais recentes primeiro)
    escalas_disponiveis = []
    if om:
        escalas_disponiveis = list(
            Escala.objects.filter(
                organizacao_militar=om,
                tipo_escala=tipo_escala,
                status__in=('publicada', 'previsao'),
            ).order_by('-ano', '-mes')
        )

    # Determina qual escala exibir
    escala_atual = None
    if om:
        if req_mes and req_ano:
            # Mês específico solicitado via ?mes=&ano=
            escala_atual = next(
                (e for e in escalas_disponiveis if e.mes == req_mes and e.ano == req_ano),
                None,
            )
        if not escala_atual:
            # Padrão: mês corrente ou o mais recente disponível
            escala_atual = next(
                (e for e in escalas_disponiveis if e.mes == hoje.month and e.ano == hoje.year),
                None,
            )
            if not escala_atual and escalas_disponiveis:
                # Fallback: mais recente não futuro
                escala_atual = next(
                    (e for e in escalas_disponiveis
                     if (e.ano, e.mes) <= (hoje.year, hoje.month)),
                    escalas_disponiveis[0] if escalas_disponiveis else None,
                )

    # Escalas anterior e próxima para os botões de navegação
    escala_anterior = None
    escala_proxima = None
    if escala_atual and escalas_disponiveis:
        idx = next(
            (i for i, e in enumerate(escalas_disponiveis)
             if e.mes == escala_atual.mes and e.ano == escala_atual.ano),
            None,
        )
        if idx is not None:
            if idx + 1 < len(escalas_disponiveis):
                escala_anterior = escalas_disponiveis[idx + 1]   # lista decrescente
            if idx > 0:
                escala_proxima = escalas_disponiveis[idx - 1]

    def _extrair_itens(escala, data_inicio=None, data_fim=None):
        """Extrai todos os itens de uma escala como lista de dicts."""
        qs = EscalaItem.objects.filter(escala=escala)
        if data_inicio:
            qs = qs.filter(calendario_dia__data__gte=data_inicio)
        if data_fim:
            qs = qs.filter(calendario_dia__data__lte=data_fim)
        qs = qs.select_related(
            'militar__posto',
            'substituto__posto',
            'calendario_dia__tipo_servico',
        ).order_by('calendario_dia__data')
        dias_map = {}
        for item in qs:
            dt = item.calendario_dia.data
            if dt not in dias_map:
                dias_map[dt] = {
                    'data': dt,
                    'dia_semana': DIAS_SEMANA[dt.weekday()],
                    'tipo_servico': item.calendario_dia.tipo_servico,
                    'militar': item.militar,
                    'substituto': item.substituto,
                }
        return [dias_map[d] for d in sorted(dias_map)]

    itens_atual = []
    proxima_escala = None
    itens_proximo = []

    if escala_atual:
        # Todos os dias do mês — sem corte por data de hoje
        ultimo_dia_mes_atual = _cal.monthrange(escala_atual.ano, escala_atual.mes)[1]
        data_inicio_atual = _d(escala_atual.ano, escala_atual.mes, 1)
        data_fim_atual = _d(escala_atual.ano, escala_atual.mes, ultimo_dia_mes_atual)
        itens_atual = _extrair_itens(escala_atual, data_inicio=data_inicio_atual, data_fim=data_fim_atual)

        # Próximo mês só aparece quando estamos vendo o mês corrente
        # (não faz sentido mostrar "próximo" ao navegar para meses passados)
        eh_mes_atual = (escala_atual.mes == hoje.month and escala_atual.ano == hoje.year)
        prox_data = data_fim_atual + _td(days=1)
        proxima_escala = (
            Escala.objects.filter(
                organizacao_militar=om,
                tipo_escala=tipo_escala,
                status__in=('previsao', 'publicada'),
                ano=prox_data.year,
                mes=prox_data.month,
            ).first()
        )

        if proxima_escala:
            # Próximo mês: todos os dias do mês completo
            ultimo_dia_prox = _cal.monthrange(proxima_escala.ano, proxima_escala.mes)[1]
            data_fim_prox = _d(proxima_escala.ano, proxima_escala.mes, ultimo_dia_prox)
            itens_proximo = _extrair_itens(
                proxima_escala,
                data_inicio=_d(proxima_escala.ano, proxima_escala.mes, 1),
                data_fim=data_fim_prox,
            )

    # Ranking "Próximos a Escalar" — usa EXATAMENTE a mesma chave de ordenação do motor.
    #
    # Ranking "Próximos a Escalar"
    #
    # Regras (idênticas ao MotorEscalaVertical):
    #   1. Ordenação: MENOR total geral (preto + vermelho + roxo) vai primeiro
    #   2. Empate no total: BASE→TOPO (mais moderno tem prioridade)
    #      lista_militares =.order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
    #      desempate_por_id = n-1-i  →  BASE (índice n-1) = 0 = prioridade máxima no empate
    #
    # Exibição: mostra a contagem de CADA tipo separado + o total geral.
    # Ano: usa o ano da escala sendo exibida (ou ano atual se não houver escala).
    ano_ranking = escala_atual.ano if escala_atual else hoje.year

    # Militares na ordem exata do motor (crítico para o desempate_por_id)
    militares_om = (
        list(
            Militar.objects.filter(organizacao_militar=om, ativo=True)
            .select_related('posto')
            .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
        )
        if om else []
    )
    n_mil = len(militares_om)
    # Chave de desempate: BASE (mais moderno, índice n-1) → 0 (prioridade mais alta no empate)
    desempate_key = {m.id: (n_mil - 1 - i) for i, m in enumerate(militares_om)}

    # Carrega todos os Quadrinhos do ano de uma vez só (1 query)
    quadrinhos_ano = Quadrinho.objects.filter(
        militar__in=militares_om,
        tipo_escala=tipo_escala,
        ano=ano_ranking,
    ).select_related('tipo_servico')

    # Monta dict: militar_id → {total_geral, por_tipo: {tipo_servico_id: count}}
    totais_map = {
        m.id: {'militar': m, 'total_geral': 0, 'por_tipo': {}}
        for m in militares_om
    }
    for qd in quadrinhos_ano:
        if qd.militar_id in totais_map:
            totais_map[qd.militar_id]['total_geral'] += qd.total
            totais_map[qd.militar_id]['por_tipo'][qd.tipo_servico_id] = qd.total

    # Adiciona lançamentos manuais (lastro, atestado, etc.) ao total geral
    # — exatamente como o motor e o quadrinho visual contabilizam
    lancamentos_pub = LancamentoManualQuadrinho.objects.filter(
        militar__in=militares_om,
        tipo_escala=tipo_escala,
        ano=ano_ranking,
    )
    for lm in lancamentos_pub:
        if lm.militar_id in totais_map:
            totais_map[lm.militar_id]['total_geral'] += lm.quantidade

    # Para cada tipo de serviço: ranking pelo total_geral, exibe contagem do tipo específico
    proximos_por_servico = []
    for ts in tipos_servico:
        candidatos = [
            {
                'militar': v['militar'],
                'total': v['total_geral'],                      # total geral (preto+verm+roxo)
                'contagem_tipo': v['por_tipo'].get(ts.id, 0),  # só este tipo
            }
            for v in totais_map.values()
        ]
        # Menor total geral → BASE→TOPO no empate (igual ao motor)
        proximos = sorted(
            candidatos,
            key=lambda x: (x['total'], desempate_key.get(x['militar'].id, 0)),
        )[:5]
        proximos_por_servico.append({'tipo_servico': ts, 'proximos': proximos})

    return render(request, 'escala/escala_publica.html', {
        'om': om,
        'hoje': hoje,
        'todos_tipos': todos_tipos,
        'tipo_escala': tipo_escala,
        'escala_atual': escala_atual,
        'itens_atual': itens_atual,
        'proxima_escala': proxima_escala,
        'itens_proximo': itens_proximo,
        'proximos_por_servico': proximos_por_servico,
        'escala_anterior': escala_anterior,
        'escala_proxima': escala_proxima,
        'escalas_disponiveis': escalas_disponiveis,
        'nomes_meses': NOMES_MESES_PT,
        'eh_mes_atual': eh_mes_atual if escala_atual else False,
    })


def matriz_publica(request, slug):
    """
    Página pública (sem login) de matriz para um tipo específico.
    URL: /matriz/<slug>/   ex: /matriz/permanencia/
    Espelha /escalas/<id>/matriz/ mas sem login e com navegação por slug.
    Gerada automaticamente para cada TipoEscala cadastrado.
    """
    import calendar as _cal
    from datetime import date as _d

    hoje = _d.today()

    tipo_escala = get_object_or_404(TipoEscala, slug=slug, ativo=True)
    om = OrganizacaoMilitar.objects.filter(ativo=True).order_by('id').first()

    todos_tipos = list(TipoEscala.objects.filter(ativo=True).order_by('nome'))

    escala = (
        Escala.objects.filter(
            organizacao_militar=om,
            tipo_escala=tipo_escala,
            status__in=('publicada', 'previsao'),
        ).order_by('-ano', '-mes').first()
        if om else None
    )

    militares = []
    dias = []
    itens = []
    matrix_rows = []
    max_eventos = 0

    if escala:
        militares = list(
            Militar.objects.filter(organizacao_militar=om, ativo=True)
            .select_related('posto')
            .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
        )

        dias = list(
            CalendarioDia.objects.filter(
                organizacao_militar=om,
                data__year=escala.ano,
                data__month=escala.mes,
            ).select_related('tipo_servico').order_by('data')
        )

        itens = list(
            escala.itens.select_related('militar__posto', 'calendario_dia__tipo_servico')
            .order_by('calendario_dia__data')
        )
        itens_map = {item.calendario_dia.data: item.militar for item in itens}

        ultimo_dia_num = _cal.monthrange(escala.ano, escala.mes)[1]
        inicio = _d(escala.ano, escala.mes, 1)
        fim = _d(escala.ano, escala.mes, ultimo_dia_num)

        indisp_map = {}
        for ind in (
            Indisponibilidade.objects.filter(
                militar__organizacao_militar=om,
                data_inicio__lte=fim,
                data_fim__gte=inicio,
            ).select_related('militar', 'tipo')
        ):
            d = ind.data_inicio
            while d <= ind.data_fim:
                if inicio <= d <= fim:
                    indisp_map[(ind.militar_id, d)] = ind.tipo.nome
                d += timedelta(days=1)

        for mil in militares:
            cells = []
            total_servicos = 0
            for dia in dias:
                d = dia.data
                serves = itens_map.get(d) == mil
                unavailable = (mil.id, d) in indisp_map
                motivo_indisp = indisp_map.get((mil.id, d), '')
                if serves:
                    total_servicos += 1
                cells.append({
                    'dia': dia,
                    'serves': serves,
                    'unavailable': unavailable,
                    'motivo': motivo_indisp,
                })
            eventos = [c for c in cells if c['serves'] or c['unavailable']]
            matrix_rows.append({
                'militar': mil,
                'cells': cells,
                'eventos': eventos,
                'total': total_servicos,
            })

        max_eventos = max((len(r['eventos']) for r in matrix_rows), default=0)
        for r in matrix_rows:
            faltam = max_eventos - len(r['eventos'])
            r['eventos_padded'] = r['eventos'] + [None] * faltam

    return render(request, 'escala/matriz_publica.html', {
        'om': om,
        'hoje': hoje,
        'todos_tipos': todos_tipos,
        'tipo_escala': tipo_escala,
        'escala': escala,
        'militares': militares,
        'itens': itens,
        'matrix_rows': matrix_rows,
        'max_eventos': max_eventos,
        'max_eventos_range': range(max_eventos),
        'nomes_meses': NOMES_MESES,
    })


# ---------------------------------------------------------------------------
# Lançamentos Manuais de Quadrinho
# ---------------------------------------------------------------------------

@login_required
def lancamento_manual_listar(request):
    om = obter_om_ativa(request)
    tipo_escala_filtro = request.GET.get('tipo_escala', '')
    tipo_servico_filtro = request.GET.get('tipo_servico', '')
    ano_atual = _date.today().year
    try:
        ano_filtro = int(request.GET.get('ano') or ano_atual)
    except ValueError:
        ano_filtro = ano_atual

    qs = LancamentoManualQuadrinho.objects.filter(
        militar__organizacao_militar=om
    ).select_related(
        'militar__posto', 'tipo_escala', 'tipo_servico'
    ).order_by('militar__posto__ordem_hierarquica', 'militar__nome_guerra', 'data_criacao')

    if tipo_escala_filtro:
        qs = qs.filter(tipo_escala_id=tipo_escala_filtro)
    if tipo_servico_filtro:
        qs = qs.filter(tipo_servico_id=tipo_servico_filtro)
    qs = qs.filter(ano=ano_filtro)

    tipos_escala = list(TipoEscala.objects.filter(ativo=True).order_by('nome'))
    tipos_servico = list(om.tipos_servico.filter(ativo=True).order_by('ordem')) if om else []
    anos = list(range(ano_atual + 1, ano_atual - 5, -1))

    return render(request, 'cadastro/lancamento_manual_list.html', {
        'lancamentos': qs,
        'om': om,
        'tipos_escala': tipos_escala,
        'tipos_servico': tipos_servico,
        'tipo_escala_filtro': tipo_escala_filtro,
        'tipo_servico_filtro': tipo_servico_filtro,
        'ano_filtro': ano_filtro,
        'anos': anos,
    })


@login_required
def lancamento_manual_form(request, lancamento_id=None):
    om = obter_om_ativa(request)
    instancia = get_object_or_404(LancamentoManualQuadrinho, pk=lancamento_id) if lancamento_id else None
    if request.method == 'POST':
        form = LancamentoManualForm(request.POST, instance=instancia, om=om)
        if form.is_valid():
            lm = form.save(commit=False)
            if not instancia:
                lm.criado_por = request.user if request.user.is_authenticated else None
            lm.save()
            acao = 'atualizado' if instancia else 'registrado'
            messages.success(request, f'Lançamento "{lm.label}" {acao} com sucesso.')
            return redirect('lancamento_manual_listar')
    else:
        form = LancamentoManualForm(instance=instancia, om=om)
    return render(request, 'cadastro/lancamento_manual_form.html', {
        'form': form,
        'lancamento': instancia,
        'om': om,
    })


@login_required
def lancamento_manual_excluir(request, lancamento_id):
    lancamento = get_object_or_404(LancamentoManualQuadrinho, pk=lancamento_id)
    if request.method == 'POST':
        label = lancamento.label
        lancamento.delete()
        messages.success(request, f'Lançamento "{label}" excluído.')
        return redirect('lancamento_manual_listar')
    return render(request, 'cadastro/lancamento_manual_confirm_delete.html', {
        'lancamento': lancamento,
    })


# ---------------------------------------------------------------------------
# Gerenciamento de Usuários (admin_om only)
# ---------------------------------------------------------------------------

@login_required
def usuario_listar(request):
    """Lista unificada de militares (= usuários) da OM, ordenada por antiguidade."""
    om = obter_om_ativa(request)
    q = request.GET.get('q', '').strip()
    perfil_filtro = request.GET.get('perfil', '')
    tipo_escala_filtro = request.GET.get('tipo_escala', '')
    divisao_filtro = request.GET.get('divisao', '')
    posto_filtro = request.GET.get('posto', '')

    # ── Militares ativos, ordem de antiguidade ────────────────────────────────
    militares_qs = Militar.objects.filter(ativo=True)
    if om:
        militares_qs = militares_qs.filter(organizacao_militar=om)
    militares_qs = militares_qs.select_related(
        'posto', 'divisao', 'especialidade', 'user', 'organizacao_militar'
    ).prefetch_related('tipos_escala')

    if q:
        militares_qs = militares_qs.filter(
            Q(nome_guerra__icontains=q) |
            Q(nome_completo__icontains=q) |
            Q(matricula__icontains=q) |
            Q(cpf__icontains=q) |
            Q(user__username__icontains=q)
        )
    if tipo_escala_filtro:
        militares_qs = militares_qs.filter(tipos_escala__id=tipo_escala_filtro)
    if perfil_filtro:
        militares_qs = militares_qs.filter(user__perfil=perfil_filtro)
    if divisao_filtro:
        militares_qs = militares_qs.filter(divisao_id=divisao_filtro)
    if posto_filtro:
        militares_qs = militares_qs.filter(posto_id=posto_filtro)

    from django.db.models import F
    militares_qs = militares_qs.order_by(
        'posto__ordem_hierarquica',
        F('data_ultima_promocao').asc(nulls_last=True),
        F('nota').desc(nulls_last=True),
        'nome_guerra',
    )

    # ── Militares inativos (sem filtros) ─────────────────────────────────────
    militares_inativos = (
        Militar.objects.filter(organizacao_militar=om, ativo=False)
        .select_related('posto', 'divisao')
        .order_by('posto__ordem_hierarquica', 'nome_guerra')
        if om else Militar.objects.none()
    )

    # ── Usuários sem militar (admins/escalantes avulsos) ─────────────────────
    usuarios_sem_militar = UsuarioCustomizado.objects.filter(
        militar__isnull=True
    ).select_related('om_principal').order_by('username')
    if q:
        usuarios_sem_militar = usuarios_sem_militar.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q)
        )
    if perfil_filtro:
        usuarios_sem_militar = usuarios_sem_militar.filter(perfil=perfil_filtro)

    tipos_escala = TipoEscala.objects.filter(ativo=True).order_by('nome')
    divisoes = (
        Divisao.objects.filter(organizacao_militar=om, ativo=True).order_by('nome')
        if om else Divisao.objects.none()
    )
    postos = Posto.objects.filter(ativo=True).order_by('ordem_hierarquica')

    return render(request, 'cadastro/usuario_list.html', {
        'militares':            militares_qs,
        'militares_inativos':   militares_inativos,
        'usuarios_sem_militar': usuarios_sem_militar,
        'q':                    q,
        'perfil_filtro':        perfil_filtro,
        'tipo_escala_filtro':   tipo_escala_filtro,
        'divisao_filtro':       divisao_filtro,
        'posto_filtro':         posto_filtro,
        'perfis':               PerfilUsuario.choices,
        'tipos_escala':         tipos_escala,
        'divisoes':             divisoes,
        'postos':               postos,
        'om':                   om,
    })


@login_required
def usuario_form(request, usuario_id=None):
    instancia = get_object_or_404(UsuarioCustomizado, pk=usuario_id) if usuario_id else None
    # Usuários vinculados a um militar são editados pela página do militar
    if instancia is not None:
        militar_vinculado = getattr(instancia, 'militar', None)
        if militar_vinculado is not None:
            return redirect('militar_editar', militar_id=militar_vinculado.id)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=instancia)
        if form.is_valid():
            usuario = form.save(commit=False)
            if not instancia:
                usuario.set_unusable_password()
            usuario.save()
            acao = 'atualizado' if instancia else 'criado'
            messages.success(request, f'Usuário {usuario.username} {acao} com sucesso.')
            return redirect('usuario_listar')
    else:
        form = UsuarioForm(instance=instancia)
    return render(request, 'cadastro/usuario_form.html', {
        'form': form,
        'usuario': instancia,
    })


@login_required
def usuario_excluir(request, usuario_id):
    usuario = get_object_or_404(UsuarioCustomizado, pk=usuario_id)
    if usuario.is_superuser:
        messages.error(request, 'Superusuários não podem ser desativados por esta tela.')
        return redirect('usuario_listar')
    if request.method == 'POST':
        usuario.ativo = False
        usuario.is_active = False
        usuario.save()
        messages.success(request, f'Usuário {usuario.username} desativado.')
        return redirect('usuario_listar')
    return render(request, 'cadastro/usuario_confirm_delete.html', {'usuario': usuario})


# ---------------------------------------------------------------------------
# Exportação Excel de Quadrinho
# ---------------------------------------------------------------------------

@login_required
def quadrinho_exportar(request):
    """Exporta o quadrinho em formato Excel com matriz de serviços."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from calendar import monthrange

    om = obter_om_ativa(request)
    if not om:
        messages.error(request, 'Nenhuma OM ativa.')
        return redirect('quadrinho_visao')

    # Parâmetros
    try:
        ano = int(request.GET.get('ano') or _date.today().year)
    except ValueError:
        ano = _date.today().year

    tipo_escala_id = request.GET.get('tipo_escala')
    tipo_escala = get_object_or_404(TipoEscala, pk=tipo_escala_id) if tipo_escala_id else TipoEscala.objects.first()

    if not tipo_escala:
        messages.error(request, 'Nenhum tipo de escala disponível.')
        return redirect('quadrinho_visao')

    # Obter militares
    militares = list(
        Militar.objects.filter(organizacao_militar=om, ativo=True)
        .select_related('posto')
        .order_by('posto__ordem_hierarquica', 'data_ultima_promocao', '-nota', 'nome_guerra')
    )

    if not militares:
        messages.error(request, 'Nenhum militar ativo nesta OM.')
        return redirect('quadrinho_visao')

    # Obter tipos de serviço
    tipos_servico = list(om.tipos_servico.filter(ativo=True).order_by('ordem'))

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Quadrinho {ano}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003A78", end_color="003A78", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Cabeçalho
    ws.merge_cells('A1:A2')
    ws['A1'] = "Militar"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    ws['A1'].border = thin_border

    ws.merge_cells('B1:B2')
    ws['B1'] = "Posto"
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    ws['B1'].alignment = center_align
    ws['B1'].border = thin_border

    # Meses como colunas (cada mês com seus dias)
    col_idx = 3  # Começa na coluna C
    meses = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
    ]

    mapa_servicos = {}
    if tipo_escala:
        servicos = EscalaItem.objects.filter(
            militar__in=militares,
            escala__tipo_escala=tipo_escala,
            calendario_dia__data__year=ano,
        ).values('militar_id', 'calendario_dia__data', 'calendario_dia__tipo_servico__cor_hex')

        def converter_cor_para_excel(cor_hex):
            """Converte cor #RRGGBB para formato aRGB do openpyxl"""
            if not cor_hex:
                return None
            cor_hex = cor_hex.lstrip('#')
            if len(cor_hex) == 6:
                return f"FF{cor_hex.upper()}"  # FF prefix = opacity 100%
            return None

        for s in servicos:
            chave = (s['militar_id'], s['calendario_dia__data'])
            cor_original = s['calendario_dia__tipo_servico__cor_hex']
            mapa_servicos[chave] = converter_cor_para_excel(cor_original)

    # Para cada mês, criar colunas de dias
    for mes_num, mes_nome in meses:
        ultimo_dia = monthrange(ano, mes_num)[1]

        # Cabeçalho do mês
        col_start = col_idx
        col_end = col_idx + ultimo_dia - 1
        if col_start == col_end:
            cell_ref = f"{get_column_letter(col_start)}1"
        else:
            cell_ref = f"{get_column_letter(col_start)}1:{get_column_letter(col_end)}1"
        ws.merge_cells(cell_ref)
        ws[f"{get_column_letter(col_start)}1"] = mes_nome
        ws[f"{get_column_letter(col_start)}1"].font = header_font
        ws[f"{get_column_letter(col_start)}1"].fill = header_fill
        ws[f"{get_column_letter(col_start)}1"].alignment = center_align

        # Dias do mês na linha 2
        for dia in range(1, ultimo_dia + 1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}2"] = dia
            ws[f"{col_letter}2"].font = Font(size=8)
            ws[f"{col_letter}2"].alignment = center_align
            ws[f"{col_letter}2"].border = thin_border
            col_idx += 1

    # Coluna Total
    total_col = col_idx
    ws.merge_cells(f"{get_column_letter(total_col)}1:{get_column_letter(total_col)}2")
    ws[f"{get_column_letter(total_col)}1"] = "Total"
    ws[f"{get_column_letter(total_col)}1"].font = header_font
    ws[f"{get_column_letter(total_col)}1"].fill = header_fill
    ws[f"{get_column_letter(total_col)}1"].alignment = center_align

    # Linha de dados
    for row_idx, militar in enumerate(militares, start=3):
        # Militar
        col_letter = 'A'
        ws[f"{col_letter}{row_idx}"] = militar.nome_guerra
        ws[f"{col_letter}{row_idx}"].border = thin_border

        # Posto
        col_letter = 'B'
        ws[f"{col_letter}{row_idx}"] = militar.posto.sigla
        ws[f"{col_letter}{row_idx}"].alignment = center_align
        ws[f"{col_letter}{row_idx}"].border = thin_border

        # Dias
        total_servicos = 0
        col_idx = 3
        for mes_num, _ in meses:
            ultimo_dia = monthrange(ano, mes_num)[1]
            for dia in range(1, ultimo_dia + 1):
                data = _date(ano, mes_num, dia)
                cor = mapa_servicos.get((militar.id, data))

                col_letter = get_column_letter(col_idx)
                if cor:
                    ws[f"{col_letter}{row_idx}"].fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                    ws[f"{col_letter}{row_idx}"].alignment = center_align
                    total_servicos += 1
                else:
                    ws[f"{col_letter}{row_idx}"].value = "·"
                    ws[f"{col_letter}{row_idx}"].alignment = Alignment(horizontal='center')
                    ws[f"{col_letter}{row_idx}"].font = Font(color="CCCCCC")

                ws[f"{col_letter}{row_idx}"].border = thin_border
                col_idx += 1

        # Total
        col_letter = get_column_letter(total_col)
        ws[f"{col_letter}{row_idx}"] = total_servicos
        ws[f"{col_letter}{row_idx}"].alignment = center_align
        ws[f"{col_letter}{row_idx}"].font = Font(bold=True)
        ws[f"{col_letter}{row_idx}"].border = thin_border

    # Ajustar larguras
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    for col in range(3, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 3

    # Resposta HTTP
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=quadinho_{ano}_{tipo_escala.slug}.xlsx'

    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# TROCA DE SERVIÇO
# ---------------------------------------------------------------------------

@login_required
def troca_listar(request):
    """Lista de trocas de serviço - diferenciado por perfil"""
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('dashboard')

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)

    # Se é militar comum, mostra só as suas próprias trocas
    if usuario.perfil == PerfilUsuario.MILITAR and militar_logado:
        # Mostrar trocas onde é quem sai OU quem entra
        trocas = TrocaServico.objects.filter(
            organizacao_militar=om
        ).filter(
            Q(militar_sai=militar_logado) |
            Q(militar_entra=militar_logado) |
            Q(militar_entra_2=militar_logado)
        ).select_related(
            'militar_sai', 'militar_entra', 'militar_sai_2', 'militar_entra_2',
            'tipo_escala', 'organizacao_militar'
        ).order_by('-data_criacao')

        # Contagem para o mês atual
        hoje = date.today()
        trocas_mes = trocas.filter(data_criacao__year=hoje.year, data_criacao__month=hoje.month).count()

        return render(request, 'troca/troca_militar_listar.html', {
            'trocas': trocas,
            'trocas_mes': trocas_mes,
            'om': om,
            'militar_logado': militar_logado,
        })

    # Escalante, Chefe, Adjunto e Admin: vêem TODAS as trocas da OM
    # Chefe/Adjunto também vêem o painel de homologação (aba extra no template)
    elif usuario.pode_gerenciar_escalas():
        trocas = TrocaServico.objects.filter(
            organizacao_militar=om
        ).select_related(
            'militar_sai', 'militar_entra', 'militar_sai_2', 'militar_entra_2',
            'tipo_escala', 'organizacao_militar'
        ).order_by('-data_criacao')

        # Trocas aguardando homologação (só relevante para chefe/adjunto)
        trocas_para_homologar = TrocaServico.objects.filter(
            organizacao_militar=om,
            status='aprovada',
            escala_status='publicada',
        ).count() if usuario.pode_publicar_escala() else 0

        pagina = int(request.GET.get('pagina', 1))
        por_pagina = 20
        total = trocas.count()
        trocas_pagina = trocas[(pagina - 1) * por_pagina:pagina * por_pagina]

        hoje = date.today()
        trocas_mes = trocas.filter(data_criacao__year=hoje.year, data_criacao__month=hoje.month).count()

        return render(request, 'troca/troca_escalante_listar.html', {
            'trocas': trocas_pagina,
            'total': total,
            'pagina': pagina,
            'por_pagina': por_pagina,
            'total_paginas': (total + por_pagina - 1) // por_pagina,
            'trocas_mes': trocas_mes,
            'trocas_para_homologar': trocas_para_homologar,
            'pode_homologar': usuario.pode_publicar_escala(),
            'om': om,
        })

    # Outros perfis (admin, gerente, etc) - redireciona para escala
    messages.info(request, "Você não tem acesso a esta funcionalidade.")
    return redirect('dashboard')


@login_required
def troca_servicos_militar(request):
    """
    Retorna os serviços futuros de um militar (hoje → próximos 3 meses).

    Parâmetros GET:
      tipo_escala  (int, opcional) — filtra por tipo de escala
      militar_id   (int, opcional) — se informado, retorna serviços desse
                                     militar em vez do logado (deve pertencer
                                     à mesma OM). Usado na segunda perna da
                                     troca mútua para exibir os dias do
                                     militar que vai entrar.
    """
    om = obter_om_ativa(request)
    if not om:
        return JsonResponse({'erro': 'Nenhuma OM ativa'}, status=400)

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)
    if not militar_logado:
        return JsonResponse({'erro': 'Usuário não é militar'}, status=400)

    tipo_escala_id = request.GET.get('tipo_escala')
    militar_id = request.GET.get('militar_id')

    # Decide qual militar consultar
    if militar_id:
        try:
            militar_alvo = Militar.objects.get(pk=militar_id, organizacao_militar=om, ativo=True)
        except Militar.DoesNotExist:
            return JsonResponse({'erro': 'Militar não encontrado nesta OM'}, status=404)
    else:
        militar_alvo = militar_logado

    hoje = date.today()
    limite = hoje + timedelta(days=92)  # próximos 3 meses

    escalas = Escala.objects.filter(
        organizacao_militar=om,
        status__in=['previsao', 'publicada'],
    )
    if tipo_escala_id:
        escalas = escalas.filter(tipo_escala_id=tipo_escala_id)

    servicos = []
    for escala in escalas:
        itens = EscalaItem.objects.filter(
            escala=escala,
            militar=militar_alvo,
            calendario_dia__data__gte=hoje,
            calendario_dia__data__lte=limite,
        ).select_related('calendario_dia', 'calendario_dia__tipo_servico')

        for item in itens:
            servicos.append({
                'id': item.id,
                'data': item.calendario_dia.data.strftime('%Y-%m-%d'),
                'data_display': item.calendario_dia.data.strftime('%d/%m/%Y'),
                'tipo_servico': item.calendario_dia.tipo_servico.nome,
                'tipo_servico_cor': item.calendario_dia.tipo_servico.cor_hex,
                'escala_tipo': escala.tipo_escala.nome,
                'escala_status': escala.status,
            })

    servicos.sort(key=lambda x: x['data'])
    return JsonResponse({'servicos': servicos})


@login_required
def troca_solicitar(request):
    """Formulário para solicitar troca de serviço"""
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('dashboard')

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)
    if not militar_logado:
        messages.error(request, "Apenas militares podem solicitar troca.")
        return redirect('dashboard')

    # Se veio do POST, processar o formulário
    if request.method == 'POST':
        tipo_troca = request.POST.get('tipo_troca', 'simples')
        tipo_escala_id = request.POST.get('tipo_escala')
        escala_status = request.POST.get('escala_status')
        motivo = request.POST.get('motivo')

        # Dados da primeira perna (militar que sai)
        data_servico_sai = request.POST.get('data_servico_sai')
        militar_entra_id = request.POST.get('militar_entra')

        if not all([tipo_escala_id, escala_status, data_servico_sai, militar_entra_id, motivo]):
            messages.error(request, "Preencha todos os campos obrigatórios.")
            return redirect('troca_solicitar')

        tipo_escala = get_object_or_404(TipoEscala, id=tipo_escala_id)
        militar_entra = get_object_or_404(Militar, id=militar_entra_id, organizacao_militar=om)

        # Criar a troca
        troca = TrocaServico.objects.create(
            organizacao_militar=om,
            tipo_escala=tipo_escala,
            tipo_troca=tipo_troca,
            escala_status=escala_status,
            militar_sai=militar_logado,
            data_servico_sai=data_servico_sai,
            militar_entra=militar_entra,
            motivo=motivo,
            aprovada_militar_sai=True,  # Quem solicita já aprova
            usuario_solicitacao=usuario,
        )

        # Se é troca mútua, criar a segunda perna
        if tipo_troca == 'mutua':
            militar_entra_2_id = request.POST.get('militar_entra_2')
            data_servico_sai_2 = request.POST.get('data_servico_sai_2')

            if not militar_entra_2_id or not data_servico_sai_2:
                messages.error(request, "Para troca mútua, preencha os dados do segundo militar.")
                troca.delete()
                return redirect('troca_solicitar')

            militar_entra_2 = get_object_or_404(Militar, id=militar_entra_2_id, organizacao_militar=om)

            # Buscar o militar que sai do segundo dia (o que o militar_logado vai assumir)
            # Na verdade, para a segunda perna:
            # - O militar_logado vai assumir o serviço do dia data_servico_sai_2 (do militar_entra_2)
            # - O militar_entra_2 vai assumir o serviço do dia data_servico_sai (do militar_logado)
            troca.militar_sai_2 = militar_entra_2     # outro militar sai do dia 2
            troca.data_servico_sai_2 = data_servico_sai_2
            troca.militar_entra_2 = militar_logado      # iniciador entra no dia 2
            troca.aprovada_militar_entra_2 = True       # iniciador já aprova a sua entrada
            # aprovada_militar_sai_2 fica None → aguardando o militar_entra_2 aceitar
            troca.save()

        messages.success(request, f"Troca solicitada! Número de controle: {troca.numero_controle}")
        return redirect('troca_listar')

    # GET - mostrar formulário
    # Buscar tipos de escala
    tipos_escala = TipoEscala.objects.filter(ativo=True).order_by('nome')

    # Buscar militares da OM (exceto o logado)
    militares = Militar.objects.filter(
        organizacao_militar=om,
        ativo=True
    ).exclude(id=militar_logado.id).select_related('posto').order_by('nome_guerra')

    return render(request, 'troca/troca_solicitar.html', {
        'tipos_escala': tipos_escala,
        'militares': militares,
        'militar_logado': militar_logado,
        'om': om,
    })


def _executar_troca_na_escala(troca, om):
    """
    Aplica a troca nos EscalaItem e atualiza Quadrinhos.
    Retorna lista de alertas (string). Lista vazia = tudo OK.
    Funciona tanto para previsão (escalante aprova) quanto para publicada (chefe homologa).
    """
    alertas = []

    # --- Primeira perna ---
    escala_1 = Escala.objects.filter(
        organizacao_militar=om,
        tipo_escala=troca.tipo_escala,
        mes=troca.data_servico_sai.month,
        ano=troca.data_servico_sai.year,
        status__in=['previsao', 'publicada'],
    ).first()

    if not escala_1:
        alertas.append(
            f"Escala não encontrada para {troca.data_servico_sai:%d/%m/%Y} "
            f"(tipo: {troca.tipo_escala.nome})."
        )
    else:
        try:
            item = EscalaItem.objects.get(
                escala=escala_1,
                calendario_dia__data=troca.data_servico_sai,
                militar=troca.militar_sai,
            )
            item.militar = troca.militar_entra
            item.observacao = (
                f"Troca {troca.numero_controle}: "
                f"{troca.militar_sai.nome_guerra} -> {troca.militar_entra.nome_guerra}"
            )
            item.save()
            Quadrinho.incrementar(
                militar=troca.militar_entra,
                tipo_escala=troca.tipo_escala,
                tipo_servico=item.calendario_dia.tipo_servico,
                ano=escala_1.ano,
            )
            Quadrinho.incrementar(
                militar=troca.militar_sai,
                tipo_escala=troca.tipo_escala,
                tipo_servico=item.calendario_dia.tipo_servico,
                ano=escala_1.ano,
                quantidade=-1,
            )
        except EscalaItem.DoesNotExist:
            alertas.append(
                f"Item de escala nao encontrado: {troca.militar_sai.nome_guerra} "
                f"em {troca.data_servico_sai:%d/%m/%Y}."
            )

    # --- Segunda perna (troca mutua) ---
    if troca.tipo_troca == 'mutua' and troca.data_servico_sai_2:
        # Segunda perna pode ser em mes/escala diferente
        escala_2 = Escala.objects.filter(
            organizacao_militar=om,
            tipo_escala=troca.tipo_escala,
            mes=troca.data_servico_sai_2.month,
            ano=troca.data_servico_sai_2.year,
            status__in=['previsao', 'publicada'],
        ).first()

        if not escala_2:
            alertas.append(
                f"Escala nao encontrada para segunda perna "
                f"{troca.data_servico_sai_2:%d/%m/%Y}."
            )
        else:
            try:
                item2 = EscalaItem.objects.get(
                    escala=escala_2,
                    calendario_dia__data=troca.data_servico_sai_2,
                    militar=troca.militar_sai_2,
                )
                item2.militar = troca.militar_entra_2
                item2.observacao = (
                    f"Troca {troca.numero_controle}: "
                    f"{troca.militar_sai_2.nome_guerra} -> {troca.militar_entra_2.nome_guerra}"
                )
                item2.save()
                Quadrinho.incrementar(
                    militar=troca.militar_entra_2,
                    tipo_escala=troca.tipo_escala,
                    tipo_servico=item2.calendario_dia.tipo_servico,
                    ano=escala_2.ano,
                )
                Quadrinho.incrementar(
                    militar=troca.militar_sai_2,
                    tipo_escala=troca.tipo_escala,
                    tipo_servico=item2.calendario_dia.tipo_servico,
                    ano=escala_2.ano,
                    quantidade=-1,
                )
            except EscalaItem.DoesNotExist:
                alertas.append(
                    f"Item de escala nao encontrado: {troca.militar_sai_2.nome_guerra} "
                    f"em {troca.data_servico_sai_2:%d/%m/%Y}."
                )

    return alertas


@login_required
def troca_aceitar(request, troca_id):
    """Aceitar ou recusar uma troca (para o militar convocado)."""
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Nenhuma OM ativa.")
        return redirect('troca_listar')

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)
    if not militar_logado:
        messages.error(request, "Apenas militares podem responder a trocas.")
        return redirect('troca_listar')

    troca = get_object_or_404(TrocaServico, id=troca_id, organizacao_militar=om)

    if request.method != 'POST':
        return redirect('troca_detalhar', troca_id=troca_id)

    if troca.status != 'pendente':
        messages.warning(request, "Esta troca não está mais pendente.")
        return redirect('troca_detalhar', troca_id=troca_id)

    acao = request.POST.get('acao')  # 'aceitar', 'rejeitar' ou 'reconsiderar'

    if militar_logado == troca.militar_entra:
        campo = 'entra'
    elif militar_logado == troca.militar_entra_2:
        campo = 'entra_2'
    elif militar_logado == troca.militar_sai_2:
        campo = 'sai_2'
    else:
        messages.error(request, "Voce nao esta envolvido nesta troca.")
        return redirect('troca_listar')

    if acao == 'aceitar':
        valor = True
        msg = f"Troca {troca.numero_controle} aceita."
        ok = True
    elif acao == 'rejeitar':
        valor = False
        msg = f"Troca {troca.numero_controle} recusada."
        ok = False
    elif acao == 'reconsiderar':
        valor = None  # volta para pendente
        msg = f"Troca {troca.numero_controle} voltou para pendente — aguardando sua decisao."
        ok = None
    else:
        messages.error(request, "Acao invalida.")
        return redirect('troca_listar')

    if campo == 'entra':
        troca.aprovada_militar_entra = valor
    elif campo == 'entra_2':
        troca.aprovada_militar_entra_2 = valor
    elif campo == 'sai_2':
        troca.aprovada_militar_sai_2 = valor

    troca.save()

    if ok is True:
        messages.success(request, msg)
    else:
        messages.warning(request, msg)
    return redirect('troca_listar')


@login_required
def troca_aprovar_escalante(request, troca_id):
    """
    Escalante aprova ou reprova uma troca pendente.

    Previsao: ao aprovar, executa a troca na escala imediatamente
              (nao precisa de homologacao do chefe).
    Publicada: ao aprovar, apenas muda status para 'aprovada';
               chefe/adjunto homologa e executa depois.
    """
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('dashboard')

    usuario = request.user
    if not usuario.pode_gerenciar_escalas():
        messages.error(request, "Apenas Escalante, Chefe ou Adjunto podem aprovar trocas.")
        return redirect('dashboard')

    troca = get_object_or_404(TrocaServico, id=troca_id, organizacao_militar=om)

    if request.method != 'POST':
        return redirect('troca_listar')

    acao = request.POST.get('acao')  # 'aprovar' ou 'reprovar'
    observacao = request.POST.get('observacao', '')

    troca.escalante_observacao = observacao
    troca.data_aprovacao_escalante = timezone.now()
    troca.usuario_escalante = usuario

    if acao == 'aprovar':
        # Verificar se todos os militares envolvidos aceitaram
        todos_aceitaram = troca.aprovada_militar_sai and troca.aprovada_militar_entra
        if troca.tipo_troca == 'mutua':
            todos_aceitaram = todos_aceitaram and troca.aprovada_militar_entra_2

        if not todos_aceitaram:
            messages.error(
                request,
                "Nao e possivel aprovar: algum militar ainda nao aceitou (ou recusou) a troca."
            )
            return redirect('troca_listar')

        # Verificar se algum envolvido rejeitou
        rejeitou = (troca.aprovada_militar_entra is False)
        if troca.tipo_troca == 'mutua':
            rejeitou = rejeitou or (troca.aprovada_militar_entra_2 is False)

        if rejeitou:
            messages.error(
                request,
                "Nao e possivel aprovar: um militar recusou a troca."
            )
            return redirect('troca_listar')

        troca.aprobada_escalante = True

        if troca.escala_status == 'previsao':
            # Previsao: escalante executa direto, sem precisar do chefe
            alertas = _executar_troca_na_escala(troca, om)
            for a in alertas:
                messages.warning(request, a)
            troca.status = 'homologada'
            troca.homologada = True
            troca.data_homologacao = timezone.now()
            troca.usuario_chefe = usuario
            msg = f"Troca {troca.numero_controle} aprovada e aplicada na escala de previsao."
        else:
            # Publicada: aguarda o chefe
            troca.status = 'aprovada'
            msg = f"Troca {troca.numero_controle} aprovada. Aguardando homologacao do Chefe/Adjunto."

    else:
        troca.aprobada_escalante = False
        troca.status = 'reprovada'
        msg = f"Troca {troca.numero_controle} reprovada."

    troca.save()
    messages.success(request, msg)
    return redirect('troca_listar')


@login_required
def troca_homologar(request, troca_id):
    """
    Chefe ou Adjunto homologa (executa) ou reprova uma troca aprovada pelo escalante.
    Somente para escalas publicadas — previsao e executada direto pelo escalante.
    """
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('dashboard')

    usuario = request.user
    if not usuario.pode_publicar_escala():
        messages.error(request, "Apenas Chefes ou Adjuntos podem homologar trocas.")
        return redirect('dashboard')

    troca = get_object_or_404(TrocaServico, id=troca_id, organizacao_militar=om)

    if request.method != 'POST':
        return redirect('troca_listar')

    acao = request.POST.get('acao')  # 'homologar' ou 'reprovar'
    observacao = request.POST.get('observacao', '')

    troca.homologacao_observacao = observacao
    troca.data_homologacao = timezone.now()
    troca.usuario_chefe = usuario

    if acao == 'homologar':
        alertas = _executar_troca_na_escala(troca, om)
        for a in alertas:
            messages.warning(request, a)
        troca.homologada = True
        troca.status = 'homologada'
        msg = f"Troca {troca.numero_controle} homologada e aplicada na escala."
    else:
        troca.homologada = False
        troca.status = 'reprovada'
        msg = f"Troca {troca.numero_controle} reprovada."

    troca.save()
    messages.success(request, msg)
    return redirect('troca_listar')


@login_required
def troca_detalhar(request, troca_id):
    """Ver detalhes de uma troca."""
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('dashboard')

    troca = get_object_or_404(TrocaServico, id=troca_id, organizacao_militar=om)

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)
    pode_cancelar = (
        troca.status == 'pendente'
        and militar_logado is not None
        and militar_logado in (troca.militar_sai, troca.militar_entra, troca.militar_entra_2)
    )
    pode_editar = (
        troca.status == 'pendente'
        and militar_logado is not None
        and militar_logado == troca.militar_sai
    )

    return render(request, 'troca/troca_detalhar.html', {
        'troca': troca,
        'om': om,
        'pode_cancelar': pode_cancelar,
        'pode_editar': pode_editar,
    })


@login_required
def troca_cancelar(request, troca_id):
    """
    Cancela (exclui) uma troca pendente.
    Permitido para qualquer militar envolvido na troca enquanto status='pendente'.
    """
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('troca_listar')

    troca = get_object_or_404(TrocaServico, id=troca_id, organizacao_militar=om)

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)

    if troca.status != 'pendente':
        messages.error(request, "So e possivel cancelar trocas com status Pendente.")
        return redirect('troca_detalhar', troca_id=troca_id)

    envolvidos = [troca.militar_sai, troca.militar_entra]
    if troca.militar_entra_2:
        envolvidos.append(troca.militar_entra_2)

    if not militar_logado or militar_logado not in envolvidos:
        messages.error(request, "Voce nao tem permissao para cancelar esta troca.")
        return redirect('troca_listar')

    if request.method == 'POST':
        numero = troca.numero_controle
        troca.delete()
        messages.success(request, f"Troca {numero} cancelada e excluida.")
        return redirect('troca_listar')

    # GET — tela de confirmação
    return render(request, 'troca/troca_cancelar_confirmar.html', {
        'troca': troca,
        'om': om,
    })


@login_required
def troca_editar(request, troca_id):
    """
    Edita o motivo de uma troca pendente.
    Apenas o solicitante (militar_sai) pode editar enquanto status='pendente'.
    """
    om = obter_om_ativa(request)
    if not om:
        messages.error(request, "Selecione uma OM primeiro.")
        return redirect('troca_listar')

    troca = get_object_or_404(TrocaServico, id=troca_id, organizacao_militar=om)

    usuario = request.user
    militar_logado = getattr(usuario, 'militar', None)

    if troca.status != 'pendente':
        messages.error(request, "So e possivel editar trocas com status Pendente.")
        return redirect('troca_detalhar', troca_id=troca_id)

    if not militar_logado or militar_logado != troca.militar_sai:
        messages.error(request, "Apenas o solicitante pode editar a troca.")
        return redirect('troca_listar')

    if request.method == 'POST':
        novo_motivo = request.POST.get('motivo', '').strip()
        if not novo_motivo:
            messages.error(request, "O motivo nao pode ficar vazio.")
        else:
            troca.motivo = novo_motivo
            troca.save()
            messages.success(request, f"Troca {troca.numero_controle} atualizada.")
            return redirect('troca_detalhar', troca_id=troca_id)

    return render(request, 'troca/troca_editar.html', {
        'troca': troca,
        'om': om,
    })
